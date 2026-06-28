# -*- coding: utf-8 -*-
"""
╔════════════════════════════════════════════════════════════════════════════╗
║  📚 DİJİTAL KÜTÜPHANE SİSTEMİ — TIER 1: PRODUCTION CODE                  ║
║  PyQt5 + SQLite3 + Dark Luxury Tema                                        ║
║  Rework: Kurus_platformu + Depo'dan Import Pattern Uyarlı                 ║
║  ⏰ TAM GERÇEK ZAMAN SENKRONIZASYONU (Real-time Clock Sync)               ║
║                                                                             ║
║  • Saat widget: Analog clock (gerçek sistem zamanı, her saniye güncelle)   ║
║  • Veritabanı: CURRENT_TIMESTAMP ile otomatik log tarihleri                ║
║  • Örnek veriler: Dinamik zamanlarla yüklenir (şu ana göre offset)        ║
║  • Gecikme kontrolleri: datetime('now') ile canlı hesaplanır              ║
║  • Audit log: Tüm işlemler gerçek zaman kaydıyla saklanır                  ║
║                                                                             ║
║  Çalıştır: python kutuphaneler_sistemi.py                                 ║
╚════════════════════════════════════════════════════════════════════════════╝
"""

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 0: IMPORT & PATH MANAGEMENT (kurs_platformu.py stili)
# ═════════════════════════════════════════════════════════════════════════════

import sys as _sys, os as _os
_this_dir = _os.path.dirname(_os.path.abspath(__file__))
while _this_dir in _sys.path:
    _sys.path.remove(_this_dir)

import sys, sqlite3, hashlib, os
from contextlib import contextmanager
from datetime import datetime, timedelta

# PyQt5 Importları (Depo'dan uyarlı - gerekli olanları al)
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QPushButton, QLineEdit, QDialog, QTabWidget, QFrame,
    QMessageBox, QTableWidget, QTableWidgetItem, QHeaderView,
    QComboBox, QTextEdit, QSpinBox, QDateEdit, QProgressBar,
    QScrollArea, QSizePolicy
)
from PyQt5.QtCore import Qt, QTimer, QDate, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QColor, QIcon

# Matplotlib (TIER 2 için hazır ama TIER 1'de kullanmayız)
try:
    import matplotlib
    matplotlib.use("Qt5Agg")
    from matplotlib.figure import Figure as MplFigure
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    MATPLOTLIB_OK = True
except ImportError:
    MATPLOTLIB_OK = False

# Excel support (TIER 3 için hazır)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 1: COLORS & THEME (Depo'dan uyarlı luxury dark)
# ═════════════════════════════════════════════════════════════════════════════

COLORS = {
    "bg_dark": "#0f0f1a",
    "bg_card": "#1a1a2e",
    "bg_surface": "#2d2d3a",
    "primary": "#0077b6",
    "primary_h": "#005a91",
    "secondary": "#00b4d8",
    "success": "#4CAF50",
    "success_h": "#388E3C",
    "warning": "#ff9f1c",
    "danger": "#e63946",
    "danger_h": "#c62828",
    "text_main": "#f4f4f5",
    "text_sec": "#a1a1aa",
    "border": "#3f3f46",
    "gold": "#ffd700",
}

LIBRARY_STYLESHEET = f"""
    QMainWindow {{
        background-color: {COLORS['bg_dark']};
    }}
    
    QWidget {{
        background-color: transparent;
    }}
    
    QTabWidget::pane {{
        border: none;
        background-color: {COLORS['bg_card']};
        border-radius: 12px;
    }}
    
    QTabBar::tab {{
        background-color: {COLORS['bg_surface']};
        color: {COLORS['text_main']};
        padding: 12px 30px;
        margin-right: 5px;
        border-top-left-radius: 10px;
        border-top-right-radius: 10px;
        font-weight: bold;
        font-size: 13px;
    }}
    
    QTabBar::tab:selected {{
        background-color: {COLORS['primary']};
        color: white;
    }}
    
    QTabBar::tab:hover:!selected {{
        background-color: {COLORS['bg_card']};
    }}
    
    QTableWidget {{
        background-color: {COLORS['bg_card']};
        alternate-background-color: {COLORS['bg_surface']};
        color: {COLORS['text_main']};
        gridline-color: {COLORS['border']};
        border: none;
        border-radius: 10px;
    }}
    
    QTableWidget::item {{
        padding: 10px;
    }}
    
    QTableWidget::item:selected {{
        background-color: {COLORS['primary']};
        color: white;
    }}
    
    QHeaderView::section {{
        background-color: {COLORS['bg_dark']};
        color: {COLORS['primary']};
        font-weight: bold;
        padding: 12px;
        border: none;
    }}
    
    QPushButton {{
        background-color: {COLORS['primary']};
        color: white;
        border: none;
        padding: 10px 25px;
        border-radius: 10px;
        font-weight: bold;
        font-size: 12px;
    }}
    
    QPushButton:hover {{
        background-color: {COLORS['primary_h']};
    }}
    
    QPushButton:pressed {{
        background-color: {COLORS['danger']};
    }}
    
    QPushButton#danger {{
        background-color: {COLORS['danger']};
    }}
    
    QPushButton#danger:hover {{
        background-color: {COLORS['danger_h']};
    }}
    
    QPushButton#success {{
        background-color: {COLORS['success']};
    }}
    
    QPushButton#success:hover {{
        background-color: {COLORS['success_h']};
    }}
    
    QLineEdit, QComboBox, QSpinBox, QDateEdit, QTextEdit {{
        background-color: {COLORS['bg_surface']};
        border: 1px solid {COLORS['border']};
        border-radius: 8px;
        padding: 10px;
        color: {COLORS['text_main']};
        font-size: 12px;
    }}
    
    QLineEdit:focus, QComboBox:focus, QSpinBox:focus, QDateEdit:focus {{
        border: 1px solid {COLORS['primary']};
    }}
    
    QLabel {{
        color: {COLORS['text_main']};
    }}
    
    QDialog {{
        background-color: {COLORS['bg_card']};
    }}
    
    QFrame {{
        background-color: {COLORS['bg_surface']};
        border-radius: 10px;
    }}
"""

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 2: DATABASE MANAGER (Context Manager pattern - Depo tarzı)
# ═════════════════════════════════════════════════════════════════════════════

class DatabaseManager:
    """
    SQLite Veritabanı Yöneticisi
    - @contextmanager ile transaction güvenliği
    - FOREIGN KEY constraints
    - CRUD operasyonları
    """

    def __init__(self, db_name="kutuphaneler.db"):
        self.db_name = db_name
        self.create_tables()

    @contextmanager
    def get_connection(self):
        """
        Context Manager ile veritabanı bağlantısı yönetimi.
        Otomatik commit/rollback ve kapatma.
        """
        conn = sqlite3.connect(self.db_name)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        try:
            yield conn
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            conn.close()

    def create_tables(self):
        """Tüm tabloları oluştur (ilk çalışmada)"""
        with self.get_connection() as conn:
            cursor = conn.cursor()

            # ─── ŞUBELER (TIER 6) ───
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS subeler (
                    sube_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sube_adi TEXT UNIQUE NOT NULL,
                    sehir TEXT NOT NULL,
                    adres TEXT NOT NULL,
                    telefon TEXT NOT NULL,
                    mudur TEXT,
                    olusturma_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    durum TEXT DEFAULT 'Aktif'
                )
            ''')

            # ─── ROL VE İZİNLER (TIER 6) ───
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS roller (
                    rol_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    rol_adi TEXT UNIQUE NOT NULL,
                    aciklama TEXT
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS rol_izinleri (
                    izin_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    rol_id INTEGER NOT NULL,
                    izin_adi TEXT NOT NULL,
                    FOREIGN KEY (rol_id) REFERENCES roller(rol_id)
                )
            ''')

            # ─── AUDIT LOG (TIER 6) ───
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS audit_log (
                    log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    kullanici_id INTEGER NOT NULL,
                    islem TEXT NOT NULL,
                    tablo TEXT,
                    kayit_id INTEGER,
                    eski_deger TEXT,
                    yeni_deger TEXT,
                    tarih TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (kullanici_id) REFERENCES sistem_kullanicilari(kullanici_id)
                )
            ''')

            # ─── KATEGORILER ───
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS kategoriler (
                    kategori_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    kategori_adi TEXT UNIQUE NOT NULL
                )
            ''')

            # ─── KİTAPLAR (TIER 6: sube_id eklendi) ───
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS kitaplar (
                    kitap_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sube_id INTEGER DEFAULT 1,
                    ad TEXT NOT NULL,
                    yazar TEXT NOT NULL,
                    kategori_id INTEGER NOT NULL,
                    yayin_yili INTEGER,
                    sayfa_sayisi INTEGER,
                    dil TEXT DEFAULT 'Türkçe',
                    durum TEXT DEFAULT 'Mevcut',
                    aciklama TEXT,
                    eklenme_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (kategori_id) REFERENCES kategoriler(kategori_id),
                    FOREIGN KEY (sube_id) REFERENCES subeler(sube_id)
                )
            ''')

            # ─── ÜYELER (TIER 6: sube_id eklendi) ───
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS uyeler (
                    uye_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sube_id INTEGER DEFAULT 1,
                    uye_no TEXT UNIQUE NOT NULL,
                    ad TEXT NOT NULL,
                    soyad TEXT NOT NULL,
                    email TEXT UNIQUE NOT NULL,
                    telefon TEXT NOT NULL,
                    adres TEXT,
                    kayit_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    durum TEXT DEFAULT 'Aktif',
                    FOREIGN KEY (sube_id) REFERENCES subeler(sube_id)
                )
            ''')

            # ─── ÖDÜNÇ (TIER 6: sube_id eklendi) ───
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS odunc (
                    odunc_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sube_id INTEGER DEFAULT 1,
                    kitap_id INTEGER NOT NULL,
                    uye_id INTEGER NOT NULL,
                    odunc_tarihi TIMESTAMP NOT NULL,
                    iade_tarihi TIMESTAMP,
                    son_tarih TIMESTAMP NOT NULL,
                    durum TEXT DEFAULT 'Ödünçte',
                    gecikme_ucreti REAL DEFAULT 0,
                    FOREIGN KEY (kitap_id) REFERENCES kitaplar(kitap_id),
                    FOREIGN KEY (uye_id) REFERENCES uyeler(uye_id),
                    FOREIGN KEY (sube_id) REFERENCES subeler(sube_id)
                )
            ''')

            # ─── SİSTEM KULLANICILARI (TIER 6: sube_id, rol_id eklendi) ───
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS sistem_kullanicilari (
                    kullanici_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sube_id INTEGER DEFAULT 1,
                    rol_id INTEGER DEFAULT 1,
                    kullanici_adi TEXT UNIQUE NOT NULL,
                    sifre TEXT NOT NULL,
                    ad TEXT NOT NULL,
                    soyad TEXT NOT NULL,
                    rol TEXT DEFAULT 'personel',
                    durum TEXT DEFAULT 'Aktif',
                    olusturma_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (sube_id) REFERENCES subeler(sube_id),
                    FOREIGN KEY (rol_id) REFERENCES roller(rol_id)
                )
            ''')

            # ─── DEFAULT ŞUBELER (TIER 6) ───
            cursor.execute(
                'SELECT COUNT(*) FROM subeler'
            )
            if cursor.fetchone()[0] == 0:
                cursor.execute('''
                    INSERT INTO subeler (sube_adi, sehir, adres, telefon, mudur)
                    VALUES 
                    ('Ana Şube', 'İstanbul', 'Merkez, İstanbul', '0212-123-4567', 'Admin'),
                    ('Şubat Şubesi', 'Ankara', 'Kızılay, Ankara', '0312-123-4567', 'Müdür'),
                    ('Üçüncü Şube', 'İzmir', 'Alsancak, İzmir', '0232-123-4567', 'Müdür')
                ''')

            # ─── DEFAULT ROLLER (TIER 6) ───
            cursor.execute(
                'SELECT COUNT(*) FROM roller'
            )
            if cursor.fetchone()[0] == 0:
                cursor.execute('''
                    INSERT INTO roller (rol_adi, aciklama)
                    VALUES 
                    ('admin', 'Tüm işlemlere erişim'),
                    ('kutuphanecu', 'Kitap ve ödünç yönetimi'),
                    ('uye', 'Üye paneli (ödünç sorgusu)')
                ''')

                # ─── ROLLER İÇİN İZİNLER ───
                admin_izinler = [
                    'kitap_ekle', 'kitap_sil', 'kitap_guncelle',
                    'uye_ekle', 'uye_sil', 'uye_guncelle',
                    'odunc_ver', 'iade_al',
                    'rapor_goster', 'anomali_goster',
                    'sube_yonet', 'kullanici_yonet', 'audit_log_goster'
                ]
                for izin in admin_izinler:
                    cursor.execute(
                        'INSERT INTO rol_izinleri (rol_id, izin_adi) VALUES (1, ?)',
                        (izin,)
                    )

                kutuphanecu_izinler = [
                    'kitap_ekle', 'kitap_guncelle',
                    'uye_ekle', 'uye_guncelle',
                    'odunc_ver', 'iade_al',
                    'rapor_goster'
                ]
                for izin in kutuphanecu_izinler:
                    cursor.execute(
                        'INSERT INTO rol_izinleri (rol_id, izin_adi) VALUES (2, ?)',
                        (izin,)
                    )

                uye_izinler = ['odunc_sorgu']
                for izin in uye_izinler:
                    cursor.execute(
                        'INSERT INTO rol_izinleri (rol_id, izin_adi) VALUES (3, ?)',
                        (izin,)
                    )

            # ─── DEFAULT KATEGORILER ───
            default_categories = [
                'Roman', 'Hikaye', 'Şiir', 'Bilim Kurgu', 'Fantastik',
                'Tarih', 'Bilim', 'Sanat', 'Felsefe', 'Psikoloji',
                'Çocuk', 'Eğitim', 'Sözlük', 'Ansiklopedi', 'Diğer'
            ]
            for kategori in default_categories:
                cursor.execute(
                    'INSERT OR IGNORE INTO kategoriler (kategori_adi) VALUES (?)',
                    (kategori,)
                )

            # ─── DEFAULT ADMIN KULLANICISI ───
            cursor.execute(
                'SELECT COUNT(*) FROM sistem_kullanicilari WHERE kullanici_adi = "kutuphane"'
            )
            if cursor.fetchone()[0] == 0:
                cursor.execute('''
                    INSERT INTO sistem_kullanicilari 
                    (sube_id, rol_id, kullanici_adi, sifre, ad, soyad, rol)
                    VALUES (1, 1, 'kutuphane', '12345', 'Admin', 'Kütüphane', 'admin')
                ''')

    # ─────────────────────────────────────────────────────────────────────────
    # KİTAP İŞLEMLERİ
    # ─────────────────────────────────────────────────────────────────────────

    def kategori_id_al(self, kategori_adi):
        """Kategori adından ID'sini al"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT kategori_id FROM kategoriler WHERE kategori_adi = ?', (kategori_adi,))
            row = cursor.fetchone()
            return row['kategori_id'] if row else None

    def kitap_ekle(self, ad, yazar, kategori_adi, yayin_yili=None, sayfa_sayisi=None, dil="Türkçe", aciklama=""):
        """Yeni kitap ekle"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            kategori_id = self.kategori_id_al(kategori_adi)
            if not kategori_id:
                raise ValueError(f"Kategori '{kategori_adi}' bulunamadı!")
            
            cursor.execute('''
                INSERT INTO kitaplar (ad, yazar, kategori_id, yayin_yili, sayfa_sayisi, dil, aciklama)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (ad, yazar, kategori_id, yayin_yili, sayfa_sayisi, dil, aciklama))
            return cursor.lastrowid

    def kitaplari_getir(self, kategori_adi=None):
        """Aktif kitapları getir (opsiyonel kategori filtresi). Pasif kitaplar gizli kalır."""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            if kategori_adi:
                cursor.execute('''
                    SELECT k.*, ktt.kategori_adi
                    FROM kitaplar k
                    JOIN kategoriler ktt ON k.kategori_id = ktt.kategori_id
                    WHERE k.durum != 'Pasif' AND ktt.kategori_adi = ?
                    ORDER BY k.ad
                ''', (kategori_adi,))
            else:
                cursor.execute('''
                    SELECT k.*, ktt.kategori_adi
                    FROM kitaplar k
                    JOIN kategoriler ktt ON k.kategori_id = ktt.kategori_id
                    WHERE k.durum != 'Pasif'
                    ORDER BY k.ad
                ''')
            
            return [dict(row) for row in cursor.fetchall()]

    def kitap_getir(self, kitap_id):
        """Tek kitap getir"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT k.*, ktt.kategori_adi
                FROM kitaplar k
                JOIN kategoriler ktt ON k.kategori_id = ktt.kategori_id
                WHERE k.kitap_id = ?
            ''', (kitap_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def kitap_guncelle(self, kitap_id, ad, yazar, kategori_adi, yayin_yili, sayfa_sayisi, dil, aciklama):
        """Kitap bilgilerini güncelle"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            kategori_id = self.kategori_id_al(kategori_adi)
            
            cursor.execute('''
                UPDATE kitaplar
                SET ad = ?, yazar = ?, kategori_id = ?, yayin_yili = ?, 
                    sayfa_sayisi = ?, dil = ?, aciklama = ?
                WHERE kitap_id = ?
            ''', (ad, yazar, kategori_id, yayin_yili, sayfa_sayisi, dil, aciklama, kitap_id))

    def kitap_sil(self, kitap_id):
        """Kitabı pasif duruma geçir (soft delete). Ödünçteyse engellenir.
        Geçmiş ödünç kayıtları korunur; kitap görünmez hale gelir."""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            # Kontrol: Ödünçte mi?
            cursor.execute(
                'SELECT COUNT(*) FROM odunc WHERE kitap_id = ? AND durum = "Ödünçte"',
                (kitap_id,)
            )
            if cursor.fetchone()[0] > 0:
                raise ValueError("Bu kitap ödünçte olduğu için silinemez!")
            
            # Soft delete: pasif duruma geçir (veriler korunur)
            cursor.execute('UPDATE kitaplar SET durum = \'Pasif\' WHERE kitap_id = ?', (kitap_id,))

    def kategorileri_getir(self):
        """Tüm kategorileri getir"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT kategori_adi FROM kategoriler ORDER BY kategori_adi')
            return [row['kategori_adi'] for row in cursor.fetchall()]

    # ─────────────────────────────────────────────────────────────────────────
    # ÜYE İŞLEMLERİ
    # ─────────────────────────────────────────────────────────────────────────

    def uye_ekle(self, uye_no, ad, soyad, email, telefon, adres=""):
        """Yeni üye ekle"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO uyeler (uye_no, ad, soyad, email, telefon, adres)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (uye_no, ad, soyad, email, telefon, adres))
            return cursor.lastrowid

    def uyeleri_getir(self):
        """Aktif üyeleri getir. Pasif üyeler gizli kalır."""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM uyeler WHERE durum != \'Pasif\' ORDER BY ad')
            return [dict(row) for row in cursor.fetchall()]

    def uye_getir(self, uye_id):
        """Tek üye getir"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM uyeler WHERE uye_id = ?', (uye_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def uye_guncelle(self, uye_id, ad, soyad, email, telefon, adres):
        """Üye bilgilerini güncelle"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE uyeler
                SET ad = ?, soyad = ?, email = ?, telefon = ?, adres = ?
                WHERE uye_id = ?
            ''', (ad, soyad, email, telefon, adres, uye_id))

    def uye_sil(self, uye_id):
        """Üyeyi pasif duruma geçir (soft delete). Aktif ödüncü varsa engellenir.
        Geçmiş ödünç kayıtları korunur; üye görünmez hale gelir."""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute(
                'SELECT COUNT(*) FROM odunc WHERE uye_id = ? AND durum = "Ödünçte"',
                (uye_id,)
            )
            if cursor.fetchone()[0] > 0:
                raise ValueError("Bu üyenin ödünçte kitabı var, önce iade alınmalı!")
            
            # Soft delete: pasif duruma geçir (veriler korunur)
            cursor.execute('UPDATE uyeler SET durum = \'Pasif\' WHERE uye_id = ?', (uye_id,))

    def son_uye_no(self):
        """Son üye numarası oluştur (LIB000001...)"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT uye_no FROM uyeler ORDER BY uye_id DESC LIMIT 1')
            row = cursor.fetchone()
            if row:
                num = int(row['uye_no'].replace('LIB', ''))
                return f"LIB{num + 1:06d}"
            return "LIB000001"

    # ─────────────────────────────────────────────────────────────────────────
    # ÖDÜNÇ İŞLEMLERİ
    # ─────────────────────────────────────────────────────────────────────────

    def odunc_ver(self, kitap_id, uye_id, gun_sayisi=14):
        """Kitap ödünç ver"""
        with self.get_connection() as conn:
            cursor = conn.cursor()

            # Kitap müsait mi?
            cursor.execute('SELECT durum FROM kitaplar WHERE kitap_id = ?', (kitap_id,))
            kitap = cursor.fetchone()
            if not kitap or kitap['durum'] != 'Mevcut':
                raise ValueError("Bu kitap müsait değil!")

            # Üye aktif mi?
            cursor.execute('SELECT durum FROM uyeler WHERE uye_id = ?', (uye_id,))
            uye = cursor.fetchone()
            if not uye or uye['durum'] != 'Aktif':
                raise ValueError("Bu üye aktif değil!")

            # Ödünç kaydı oluştur
            odunc_tarihi = datetime.now()
            son_tarih = odunc_tarihi + timedelta(days=gun_sayisi)

            cursor.execute('''
                INSERT INTO odunc (kitap_id, uye_id, odunc_tarihi, son_tarih, durum)
                VALUES (?, ?, ?, ?, 'Ödünçte')
            ''', (kitap_id, uye_id, odunc_tarihi, son_tarih))

            # Kitap durumunu güncelle
            cursor.execute('UPDATE kitaplar SET durum = "Ödünçte" WHERE kitap_id = ?', (kitap_id,))

    def iade_al(self, odunc_id):
        """Kitap iade al"""
        with self.get_connection() as conn:
            cursor = conn.cursor()

            cursor.execute(
                'SELECT kitap_id, son_tarih FROM odunc WHERE odunc_id = ? AND durum = "Ödünçte"',
                (odunc_id,)
            )
            odunc = cursor.fetchone()
            if not odunc:
                raise ValueError("Geçersiz ödünç kaydı!")

            iade_tarihi = datetime.now()
            son_tarih = datetime.fromisoformat(odunc['son_tarih'].replace(' ', 'T'))

            # Gecikme ücreti hesapla (5 TL/gün)
            gecikme_ucreti = 0
            if iade_tarihi > son_tarih:
                gecikme_gunu = (iade_tarihi - son_tarih).days
                gecikme_ucreti = gecikme_gunu * 5

            cursor.execute('''
                UPDATE odunc
                SET iade_tarihi = ?, durum = 'İade Edildi', gecikme_ucreti = ?
                WHERE odunc_id = ?
            ''', (iade_tarihi, gecikme_ucreti, odunc_id))

            cursor.execute(
                'UPDATE kitaplar SET durum = "Mevcut" WHERE kitap_id = ?',
                (odunc['kitap_id'],)
            )

            return gecikme_ucreti

    def odunc_listele(self, durum=None):
        """Ödünç listesini getir"""
        with self.get_connection() as conn:
            cursor = conn.cursor()

            if durum:
                cursor.execute('''
                    SELECT o.*, k.ad as kitap_adi, k.yazar, u.ad as uye_adi, u.soyad as uye_soyad, u.uye_no
                    FROM odunc o
                    JOIN kitaplar k ON o.kitap_id = k.kitap_id
                    JOIN uyeler u ON o.uye_id = u.uye_id
                    WHERE o.durum = ?
                    ORDER BY o.odunc_tarihi DESC
                ''', (durum,))
            else:
                cursor.execute('''
                    SELECT o.*, k.ad as kitap_adi, k.yazar, u.ad as uye_adi, u.soyad as uye_soyad, u.uye_no
                    FROM odunc o
                    JOIN kitaplar k ON o.kitap_id = k.kitap_id
                    JOIN uyeler u ON o.uye_id = u.uye_id
                    ORDER BY o.odunc_tarihi DESC
                ''')

            return [dict(row) for row in cursor.fetchall()]

    # ─────────────────────────────────────────────────────────────────────────
    # İSTATİSTİKLER (TIER 2)
    # ─────────────────────────────────────────────────────────────────────────

    def toplam_kitap_sayisi(self):
        """Aktif kitap sayısı (pasif hariç)"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM kitaplar WHERE durum != \'Pasif\'')
            return cursor.fetchone()[0]

    def mevcut_kitap_sayisi(self):
        """Mevcut kitap sayısı"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM kitaplar WHERE durum = "Mevcut"')
            return cursor.fetchone()[0]

    def oduncte_kitap_sayisi(self):
        """Ödünçte kitap sayısı"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM kitaplar WHERE durum = "Ödünçte"')
            return cursor.fetchone()[0]

    def toplam_uye_sayisi(self):
        """Aktif üye sayısı (pasif hariç)"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM uyeler WHERE durum != \'Pasif\'')
            return cursor.fetchone()[0]

    def aktif_odunc_sayisi(self):
        """Aktif ödünç sayısı"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM odunc WHERE durum = "Ödünçte"')
            return cursor.fetchone()[0]

    def toplam_gecikme_ucreti(self):
        """Toplam gecikme ücreti"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT SUM(gecikme_ucreti) FROM odunc')
            result = cursor.fetchone()[0]
            return result or 0

    def kategori_dagilimi(self):
        """Kategoriye göre kitap dağılımı"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT ktt.kategori_adi, COUNT(*) as sayi
                FROM kitaplar k
                JOIN kategoriler ktt ON k.kategori_id = ktt.kategori_id
                GROUP BY k.kategori_id
                ORDER BY sayi DESC
            ''')
            return [dict(row) for row in cursor.fetchall()]

    def en_cok_okunan_kitaplar(self, limit=10):
        """En çok ödünç alınan kitaplar (aktif, pasif hariç)"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT k.kitap_id, k.ad, k.yazar, COUNT(o.odunc_id) as odunc_sayisi
                FROM kitaplar k
                LEFT JOIN odunc o ON k.kitap_id = o.kitap_id
                WHERE k.durum != 'Pasif'
                GROUP BY k.kitap_id
                ORDER BY odunc_sayisi DESC
                LIMIT ?
            ''', (limit,))
            return [dict(row) for row in cursor.fetchall()]

    # ─────────────────────────────────────────────────────────────────────────
    # GAMIFICATION (TIER 3)
    # ─────────────────────────────────────────────────────────────────────────

    def gamification_table_create(self):
        """Gamification tablosunu oluştur"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS uye_gamification (
                    uye_id INTEGER PRIMARY KEY,
                    toplam_xp INTEGER DEFAULT 0,
                    seviye TEXT DEFAULT 'Bronze',
                    FOREIGN KEY (uye_id) REFERENCES uyeler(uye_id)
                )
            ''')

    def uye_xp_ekle(self, uye_id, xp):
        """Üyeye XP ekle"""
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    'SELECT toplam_xp FROM uye_gamification WHERE uye_id = ?',
                    (uye_id,)
                )
                row = cursor.fetchone()
                
                if row:
                    yeni_xp = row['toplam_xp'] + xp
                    cursor.execute(
                        'UPDATE uye_gamification SET toplam_xp = ? WHERE uye_id = ?',
                        (yeni_xp, uye_id)
                    )
                else:
                    cursor.execute(
                        'INSERT INTO uye_gamification (uye_id, toplam_xp) VALUES (?, ?)',
                        (uye_id, xp)
                    )
        except:
            pass  # Tablo yoksa atla

    def seviye_hesapla(self, xp):
        """XP'ye göre seviye hesapla"""
        if xp < 100:
            return ("🥉 Bronze", xp)
        elif xp < 500:
            return ("🥈 Silver", xp)
        else:
            return ("🥇 Gold", xp)

    def uye_seviyesi_getir(self, uye_id):
        """Üyenin seviyesini getir"""
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    'SELECT toplam_xp FROM uye_gamification WHERE uye_id = ?',
                    (uye_id,)
                )
                row = cursor.fetchone()
                if row:
                    seviye, xp = self.seviye_hesapla(row['toplam_xp'])
                    return f"{seviye} ({xp} XP)"
                return "🥉 Bronze (0 XP)"
        except:
            return "🥉 Bronze (0 XP)"

    # ─────────────────────────────────────────────────────────────────────────
    # RAPORLAR (TIER 3)
    # ─────────────────────────────────────────────────────────────────────────

    def excel_raporu_uret(self):
        """Excel raporu oluştur"""
        if not OPENPYXL_OK:
            return None
        
        from datetime import datetime as dt
        
        wb = openpyxl.Workbook()
        
        # ─── KİTAPLAR SAYFASI ───
        ws_kitap = wb.active
        ws_kitap.title = "Kitaplar"
        
        headers = ["ID", "Adı", "Yazar", "Kategori", "Yıl", "Sayfa", "Durum"]
        for col, header in enumerate(headers, 1):
            cell = ws_kitap.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0077b6", end_color="0077b6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, kitap in enumerate(self.kitaplari_getir(), 2):
            ws_kitap.cell(row=row_idx, column=1, value=kitap['kitap_id'])
            ws_kitap.cell(row=row_idx, column=2, value=kitap['ad'])
            ws_kitap.cell(row=row_idx, column=3, value=kitap['yazar'])
            ws_kitap.cell(row=row_idx, column=4, value=kitap['kategori_adi'])
            ws_kitap.cell(row=row_idx, column=5, value=kitap['yayin_yili'])
            ws_kitap.cell(row=row_idx, column=6, value=kitap['sayfa_sayisi'])
            ws_kitap.cell(row=row_idx, column=7, value=kitap['durum'])
        
        ws_kitap.column_dimensions['B'].width = 30
        ws_kitap.column_dimensions['C'].width = 20

        # ─── ÜYELER SAYFASI ───
        ws_uye = wb.create_sheet("Üyeler")
        
        headers = ["ID", "Üye No", "Ad", "Soyad", "Email", "Telefon", "Durum"]
        for col, header in enumerate(headers, 1):
            cell = ws_uye.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, uye in enumerate(self.uyeleri_getir(), 2):
            ws_uye.cell(row=row_idx, column=1, value=uye['uye_id'])
            ws_uye.cell(row=row_idx, column=2, value=uye['uye_no'])
            ws_uye.cell(row=row_idx, column=3, value=uye['ad'])
            ws_uye.cell(row=row_idx, column=4, value=uye['soyad'])
            ws_uye.cell(row=row_idx, column=5, value=uye['email'])
            ws_uye.cell(row=row_idx, column=6, value=uye['telefon'])
            ws_uye.cell(row=row_idx, column=7, value=uye['durum'])
        
        ws_uye.column_dimensions['E'].width = 25

        # ─── ÖDÜNÇ SAYFASI ───
        ws_odunc = wb.create_sheet("Ödünçler")
        
        headers = ["ID", "Kitap", "Üye", "Ödünç", "Son Tarih", "İade", "Durum", "Gecikme"]
        for col, header in enumerate(headers, 1):
            cell = ws_odunc.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="ff9f1c", end_color="ff9f1c", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, odunc in enumerate(self.odunc_listele(), 2):
            ws_odunc.cell(row=row_idx, column=1, value=odunc['odunc_id'])
            ws_odunc.cell(row=row_idx, column=2, value=odunc['kitap_adi'])
            ws_odunc.cell(row=row_idx, column=3, value=f"{odunc['uye_adi']} {odunc['uye_soyad']}")
            
            odunc_tarihi = datetime.fromisoformat(odunc['odunc_tarihi'].replace(' ', 'T'))
            ws_odunc.cell(row=row_idx, column=4, value=odunc_tarihi.strftime("%d.%m.%Y"))
            
            son_tarih = datetime.fromisoformat(odunc['son_tarih'].replace(' ', 'T'))
            ws_odunc.cell(row=row_idx, column=5, value=son_tarih.strftime("%d.%m.%Y"))
            
            iade_tarihi = datetime.fromisoformat(odunc['iade_tarihi'].replace(' ', 'T')) if odunc['iade_tarihi'] else None
            ws_odunc.cell(row=row_idx, column=6, value=iade_tarihi.strftime("%d.%m.%Y") if iade_tarihi else "-")
            
            ws_odunc.cell(row=row_idx, column=7, value=odunc['durum'])
            ws_odunc.cell(row=row_idx, column=8, value=f"{odunc['gecikme_ucreti']:.2f} TL")
        
        ws_odunc.column_dimensions['B'].width = 20
        ws_odunc.column_dimensions['C'].width = 20

        # ─── İSTATİSTİK SAYFASI ───
        ws_stat = wb.create_sheet("İstatistikler")
        
        stats = [
            ("Toplam Kitap", self.toplam_kitap_sayisi()),
            ("Mevcut Kitap", self.mevcut_kitap_sayisi()),
            ("Ödünçte Kitap", self.oduncte_kitap_sayisi()),
            ("Toplam Üye", self.toplam_uye_sayisi()),
            ("Aktif Ödünç", self.aktif_odunc_sayisi()),
            ("Toplam Gecikme Ücreti (TL)", f"{self.toplam_gecikme_ucreti():.2f}"),
        ]
        
        ws_stat.cell(row=1, column=1, value="Metrik").font = Font(bold=True)
        ws_stat.cell(row=1, column=2, value="Değer").font = Font(bold=True)
        
        for row_idx, (label, value) in enumerate(stats, 2):
            ws_stat.cell(row=row_idx, column=1, value=label)
            ws_stat.cell(row=row_idx, column=2, value=value)
        
        ws_stat.column_dimensions['A'].width = 30
        ws_stat.column_dimensions['B'].width = 20

        # Dosya kaydet
        filename = f"kutuphaneler_raporu_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        return filename

    def csv_raporu_uret(self, rapor_tipi="kitaplar"):
        """CSV raporu oluştur"""
        import csv
        from datetime import datetime as dt
        
        filename = f"kutuphaneler_{rapor_tipi}_{dt.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        if rapor_tipi == "kitaplar":
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=['kitap_id', 'ad', 'yazar', 'kategori_adi', 'durum'])
                writer.writeheader()
                writer.writerows(self.kitaplari_getir())
        
        elif rapor_tipi == "uyeler":
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=['uye_id', 'uye_no', 'ad', 'soyad', 'email', 'durum'])
                writer.writeheader()
                writer.writerows(self.uyeleri_getir())
        
        elif rapor_tipi == "odunc":
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=['odunc_id', 'kitap_adi', 'uye_adi', 'durum', 'gecikme_ucreti'])
                writer.writeheader()
                for odunc in self.odunc_listele():
                    writer.writerow({
                        'odunc_id': odunc['odunc_id'],
                        'kitap_adi': odunc['kitap_adi'],
                        'uye_adi': f"{odunc['uye_adi']} {odunc['uye_soyad']}",
                        'durum': odunc['durum'],
                        'gecikme_ucreti': f"{odunc['gecikme_ucreti']:.2f}"
                    })
        
        return filename

    # ─────────────────────────────────────────────────────────────────────────
    # ANOMALI TESPİTİ (TIER 4)
    # ─────────────────────────────────────────────────────────────────────────

    def unormal_iade_hizi_tespit(self):
        """Çok hızlı iade etme (1-2 gün) - Hırsızlık şüphesi"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT o.uye_id, u.ad, u.soyad, COUNT(*) as hizli_iade_sayisi
                FROM odunc o
                JOIN uyeler u ON o.uye_id = u.uye_id
                WHERE (julianday(o.iade_tarihi) - julianday(o.odunc_tarihi)) <= 2
                GROUP BY o.uye_id
                HAVING COUNT(*) >= 3
            ''')
            return [dict(row) for row in cursor.fetchall()]

    def sik_gecikme_yapanlar(self):
        """Sık sık gecikme yapan üyeler"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT u.uye_id, u.ad, u.soyad, u.uye_no, COUNT(*) as gecikme_sayisi,
                       SUM(o.gecikme_ucreti) as toplam_gecikme
                FROM odunc o
                JOIN uyeler u ON o.uye_id = u.uye_id
                WHERE o.gecikme_ucreti > 0
                GROUP BY o.uye_id
                HAVING COUNT(*) >= 2
                ORDER BY gecikme_sayisi DESC
            ''')
            return [dict(row) for row in cursor.fetchall()]

    def az_okunan_kitaplar(self):
        """Hiç ödünç alınmayan aktif kitaplar (pasif hariç)"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT k.kitap_id, k.ad, k.yazar, COUNT(o.odunc_id) as odunc_sayisi
                FROM kitaplar k
                LEFT JOIN odunc o ON k.kitap_id = o.kitap_id
                WHERE k.durum != 'Pasif'
                GROUP BY k.kitap_id
                HAVING COUNT(o.odunc_id) = 0
            ''')
            return [dict(row) for row in cursor.fetchall()]

    def gecikme_analizi(self):
        """Gecikmiş kitapları detaylı analiz et"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT o.*, k.ad as kitap_adi, u.ad as uye_adi, u.soyad as uye_soyad, u.telefon,
                       CAST((julianday('now') - julianday(o.son_tarih)) AS INTEGER) as gecikme_gunu
                FROM odunc o
                JOIN kitaplar k ON o.kitap_id = k.kitap_id
                JOIN uyeler u ON o.uye_id = u.uye_id
                WHERE o.durum = 'Ödünçte' AND o.son_tarih < datetime('now')
                ORDER BY gecikme_gunu DESC
            ''')
            return [dict(row) for row in cursor.fetchall()]

    # ─────────────────────────────────────────────────────────────────────────
    # SİSTEM KULLANICILARI
    # ─────────────────────────────────────────────────────────────────────────

    def ornek_veri_yukle(self, zorla=False):
        """
        Demo/örnek veri yükler (kitaplar, üyeler, ödünç geçmişi, gecikmeler).
        ⏰ TÜM TARİHLER GERÇEK SİSTEM ZAMANI İLE DİNAMİK HESAPLANIR:
           • Geçmiş ödünçler: 25-150 gün öncesi
           • Aktif ödünçler: 1-30 gün öncesi  
           • Gecikme kontrolleri: datetime('now') ile real-time çalışır
        Varsayılan: veritabanı BOŞSA yükler. Doluysa None döner (zorla=True ile zorlanır).
        Geriye özet sözlüğü döner: {kitap, uye, odunc, gecikme}.
        """
        import random
        from datetime import datetime, timedelta

        if not zorla and self.toplam_kitap_sayisi() > 0:
            return None  # zaten veri var

        random.seed(42)
        now = datetime.now()
        def ts(d):
            return d.strftime("%Y-%m-%d %H:%M:%S")

        # ── KİTAPLAR (ad, yazar, kategori, yıl, sayfa) ──
        KITAPLAR = [
            ("Suç ve Ceza", "Fyodor Dostoyevski", "Roman", 1866, 671),
            ("Sefiller", "Victor Hugo", "Roman", 1862, 1463),
            ("Simyacı", "Paulo Coelho", "Roman", 1988, 188),
            ("Tutunamayanlar", "Oğuz Atay", "Roman", 1972, 724),
            ("Kürk Mantolu Madonna", "Sabahattin Ali", "Roman", 1943, 160),
            ("Beyaz Geceler", "Fyodor Dostoyevski", "Hikaye", 1848, 96),
            ("Semaver", "Sait Faik", "Hikaye", 1936, 120),
            ("Otuz Beş Yaş", "Cahit Sıtkı Tarancı", "Şiir", 1946, 88),
            ("Dune", "Frank Herbert", "Bilim Kurgu", 1965, 412),
            ("Vakıf", "Isaac Asimov", "Bilim Kurgu", 1951, 244),
            ("1984", "George Orwell", "Bilim Kurgu", 1949, 328),
            ("Fahrenheit 451", "Ray Bradbury", "Bilim Kurgu", 1953, 256),
            ("Yüzüklerin Efendisi", "J.R.R. Tolkien", "Fantastik", 1954, 1178),
            ("Hobbit", "J.R.R. Tolkien", "Fantastik", 1937, 310),
            ("Harry Potter ve Felsefe Taşı", "J.K. Rowling", "Fantastik", 1997, 332),
            ("Nutuk", "M. Kemal Atatürk", "Tarih", 1927, 600),
            ("Sapiens", "Yuval Noah Harari", "Tarih", 2011, 443),
            ("İstanbul Hatırası", "Ahmet Ümit", "Tarih", 2010, 632),
            ("Kozmos", "Carl Sagan", "Bilim", 1980, 396),
            ("Zamanın Kısa Tarihi", "Stephen Hawking", "Bilim", 1988, 256),
            ("Bir Beyin Cerrahının Notları", "Henry Marsh", "Bilim", 2014, 288),
            ("Sanatın Öyküsü", "E. H. Gombrich", "Sanat", 1950, 688),
            ("Böyle Buyurdu Zerdüşt", "Friedrich Nietzsche", "Felsefe", 1883, 352),
            ("Devlet", "Platon", "Felsefe", -380, 416),
            ("Bilinçaltının Gücü", "Joseph Murphy", "Psikoloji", 1963, 296),
            ("Akış", "M. Csikszentmihalyi", "Psikoloji", 1990, 320),
            ("Küçük Prens", "Antoine de Saint-Exupery", "Çocuk", 1943, 96),
            ("Pal Sokağı Çocukları", "Ferenc Molnár", "Çocuk", 1906, 224),
            ("Etkili İnsanların 7 Alışkanlığı", "Stephen Covey", "Eğitim", 1989, 381),
            ("Atomic Habits", "James Clear", "Eğitim", 2018, 320),
        ]
        kitap_id = {}
        for ad, yazar, kat, yil, sayfa in KITAPLAR:
            try:
                kitap_id[ad] = self.kitap_ekle(ad, yazar, kat, yayin_yili=yil, sayfa_sayisi=sayfa)
            except Exception:
                pass

        # ── ÜYELER (ad, soyad, email) ──
        UYELER = [
            ("Ahmet", "Kaya", "ahmet.kaya@mail.com"),
            ("Zeynep", "Şahin", "zeynep.sahin@mail.com"),
            ("Mehmet", "Demir", "mehmet.demir@mail.com"),
            ("Elif", "Yıldız", "elif.yildiz@mail.com"),
            ("Can", "Aydın", "can.aydin@mail.com"),
            ("Selin", "Koç", "selin.koc@mail.com"),
            ("Burak", "Arslan", "burak.arslan@mail.com"),
            ("Deniz", "Öztürk", "deniz.ozturk@mail.com"),
            ("Ayşe", "Çelik", "ayse.celik@mail.com"),
            ("Emre", "Doğan", "emre.dogan@mail.com"),
            ("Merve", "Kurt", "merve.kurt@mail.com"),
            ("Onur", "Aksoy", "onur.aksoy@mail.com"),
            ("Buse", "Polat", "buse.polat@mail.com"),
            ("Kerem", "Erdoğan", "kerem.erdogan@mail.com"),
        ]
        ilceler = ["Kadıköy", "Beşiktaş", "Çankaya", "Konak", "Nilüfer"]
        uye_id = []
        for i, (ad, soyad, email) in enumerate(UYELER, 1):
            try:
                uye_id.append(self.uye_ekle(
                    f"UYE{i:04d}", ad, soyad, email,
                    f"05{random.randint(300000000, 599999999)}",
                    adres=random.choice(ilceler)))
            except Exception:
                pass

        # ── ÖDÜNÇ GEÇMİŞİ + AKTİF ÖDÜNÇLER ──
        populer = {
            "Suç ve Ceza": 9, "1984": 8, "Simyacı": 7, "Sapiens": 6,
            "Dune": 5, "Yüzüklerin Efendisi": 5, "Küçük Prens": 4,
            "Hobbit": 3, "Kozmos": 3, "Atomic Habits": 2,
        }
        aktif_oduncte = ["Vakıf", "Nutuk", "Devlet", "Akış",
                         "Harry Potter ve Felsefe Taşı", "Fahrenheit 451", "Sanatın Öyküsü"]
        # Sık geciken üyeler (yeterli üye varsa)
        sik_gecikenler = [uye_id[k] for k in (1, 4, 8) if k < len(uye_id)]

        if uye_id:
            with self.get_connection() as conn:
                cur = conn.cursor()
                # 1) Zamanında iade edilmiş ödünçler (popülerlik)
                for ad, adet in populer.items():
                    kid = kitap_id.get(ad)
                    if not kid:
                        continue
                    for _ in range(adet):
                        uid = random.choice(uye_id)
                        o = now - timedelta(days=random.randint(25, 150))
                        son = o + timedelta(days=14)
                        iade = son - timedelta(days=random.randint(1, 7))
                        cur.execute(
                            "INSERT INTO odunc (sube_id, kitap_id, uye_id, odunc_tarihi, son_tarih, iade_tarihi, durum, gecikme_ucreti) "
                            "VALUES (1, ?, ?, ?, ?, ?, 'İade Edildi', 0)",
                            (kid, uid, ts(o), ts(son), ts(iade)))
                # 2) Geç iade edilmiş ödünçler (sık gecikenlere 3'er — gecikme ücreti)
                populer_adlar = [a for a in populer if kitap_id.get(a)]
                for uid in sik_gecikenler:
                    for _ in range(3):
                        kid = kitap_id[random.choice(populer_adlar)]
                        o = now - timedelta(days=random.randint(30, 100))
                        son = o + timedelta(days=14)
                        gun = random.randint(3, 15)
                        iade = son + timedelta(days=gun)
                        cur.execute(
                            "INSERT INTO odunc (sube_id, kitap_id, uye_id, odunc_tarihi, son_tarih, iade_tarihi, durum, gecikme_ucreti) "
                            "VALUES (1, ?, ?, ?, ?, ?, 'İade Edildi', ?)",
                            (kid, uid, ts(o), ts(son), ts(iade), gun * 5))
                # 3) Aktif ödünçler (yarısı gecikmiş)
                for i, ad in enumerate(aktif_oduncte):
                    kid = kitap_id.get(ad)
                    if not kid:
                        continue
                    uid = random.choice(uye_id)
                    if i % 2 == 0:
                        o = now - timedelta(days=random.randint(18, 30))   # gecikmiş aktif
                    else:
                        o = now - timedelta(days=random.randint(1, 6))     # güncel aktif
                    son = o + timedelta(days=14)
                    cur.execute(
                        "INSERT INTO odunc (sube_id, kitap_id, uye_id, odunc_tarihi, son_tarih, durum, gecikme_ucreti) "
                        "VALUES (1, ?, ?, ?, ?, 'Ödünçte', 0)",
                        (kid, uid, ts(o), ts(son)))
                    cur.execute("UPDATE kitaplar SET durum='Ödünçte' WHERE kitap_id=?", (kid,))

        return {
            "kitap": self.toplam_kitap_sayisi(),
            "uye": self.toplam_uye_sayisi(),
            "odunc": self.aktif_odunc_sayisi(),
            "gecikme": self.toplam_gecikme_ucreti(),
        }

    def kullanici_kontrol(self, kullanici_adi, sifre):
        """Login kontrolü"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT * FROM sistem_kullanicilari
                WHERE kullanici_adi = ? AND sifre = ? AND durum = 'Aktif'
            ''', (kullanici_adi, sifre))
            result = cursor.fetchone()
            return dict(result) if result else None

    # ─────────────────────────────────────────────────────────────────────────
    # ŞUBELER YÖNETİMİ (TIER 6)
    # ─────────────────────────────────────────────────────────────────────────

    def subeleri_getir(self):
        """Tüm şubeleri getir"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM subeler WHERE durum = "Aktif" ORDER BY sube_adi')
            return [dict(row) for row in cursor.fetchall()]

    def sube_ekle(self, sube_adi, sehir, adres, telefon, mudur):
        """Yeni şube ekle"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO subeler (sube_adi, sehir, adres, telefon, mudur)
                VALUES (?, ?, ?, ?, ?)
            ''', (sube_adi, sehir, adres, telefon, mudur))
            return cursor.lastrowid

    def sube_bilgisi(self, sube_id):
        """Şube bilgisi getir"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM subeler WHERE sube_id = ?', (sube_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def sube_istatistikleri(self, sube_id):
        """Şubeye ait istatistikler"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            # Kitap sayısı
            cursor.execute('SELECT COUNT(*) FROM kitaplar WHERE sube_id = ?', (sube_id,))
            toplam_kitap = cursor.fetchone()[0]
            
            # Üye sayısı
            cursor.execute('SELECT COUNT(*) FROM uyeler WHERE sube_id = ?', (sube_id,))
            toplam_uye = cursor.fetchone()[0]
            
            # Ödünçte kitap
            cursor.execute('SELECT COUNT(*) FROM odunc WHERE sube_id = ? AND durum = "Ödünçte"', (sube_id,))
            oduncte = cursor.fetchone()[0]
            
            return {
                'toplam_kitap': toplam_kitap,
                'toplam_uye': toplam_uye,
                'oduncte': oduncte
            }

    # ─────────────────────────────────────────────────────────────────────────
    # AUDIT LOG (TIER 6)
    # ─────────────────────────────────────────────────────────────────────────

    def audit_log_kaydet(self, kullanici_id, islem, tablo, kayit_id, eski_deger=None, yeni_deger=None):
        """Audit log kayıt yaz (tarih otomatik: gerçek sistem zamanı)"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO audit_log 
                (kullanici_id, islem, tablo, kayit_id, eski_deger, yeni_deger, tarih)
                VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''', (kullanici_id, islem, tablo, kayit_id, eski_deger, yeni_deger))

    def audit_log_getir(self, limit=100):
        """Audit logları getir"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT a.*, k.ad as kullanici_adi, k.soyad as kullanici_soyad
                FROM audit_log a
                JOIN sistem_kullanicilari k ON a.kullanici_id = k.kullanici_id
                ORDER BY a.tarih DESC
                LIMIT ?
            ''', (limit,))
            return [dict(row) for row in cursor.fetchall()]

    def audit_log_sil_arsa(self, gun_sayisi=30):
        """Eski audit logları sil (n gün öncesi)"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                DELETE FROM audit_log
                WHERE tarih < datetime('now', '-' || ? || ' days')
            ''', (gun_sayisi,))

    # ─────────────────────────────────────────────────────────────────────────
    # İZİN KONTROL (TIER 6)
    # ─────────────────────────────────────────────────────────────────────────

    def izin_kontrol(self, rol_id, izin_adi):
        """Belirli rolün verilen izine sahip olup olmadığını kontrol et"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT COUNT(*) FROM rol_izinleri
                WHERE rol_id = ? AND izin_adi = ?
            ''', (rol_id, izin_adi))
            return cursor.fetchone()[0] > 0

    def roller_listele(self):
        """Tüm rolleri getir"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM roller')
            return [dict(row) for row in cursor.fetchall()]

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 3: LOGIN DIALOG
# ═════════════════════════════════════════════════════════════════════════════

class LoginDialog(QDialog):
    """Giriş Penceresi"""

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.kullanici = None
        self.setWindowTitle("📚 Dijital Kütüphane - Giriş")
        self.setGeometry(400, 300, 450, 350)
        self.setStyleSheet(LIBRARY_STYLESHEET)
        self.setModal(True)
        self.init_ui()

    def init_ui(self):
        """UI'yi oluştur"""
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(40, 50, 40, 40)

        # ─── BAŞLIK ───
        baslik = QLabel("📚 DİJİTAL KÜTÜPHANE")
        baslik_font = QFont("Arial", 20, QFont.Bold)
        baslik.setFont(baslik_font)
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet(f"color: {COLORS['primary']}; margin-bottom: 10px;")
        layout.addWidget(baslik)

        alt_baslik = QLabel("Sisteme Giriş Yapın")
        alt_baslik.setFont(QFont("Arial", 12))
        alt_baslik.setAlignment(Qt.AlignCenter)
        alt_baslik.setStyleSheet(f"color: {COLORS['text_sec']}; margin-bottom: 30px;")
        layout.addWidget(alt_baslik)

        # ─── KULLANICI ADI ───
        kadi_label = QLabel("Kullanıcı Adı:")
        kadi_label.setStyleSheet(f"color: {COLORS['primary']}; font-weight: bold;")
        self.kadi_input = QLineEdit()
        self.kadi_input.setPlaceholderText("kutuphane")
        layout.addWidget(kadi_label)
        layout.addWidget(self.kadi_input)

        # ─── ŞİFRE ───
        sifre_label = QLabel("Şifre:")
        sifre_label.setStyleSheet(f"color: {COLORS['primary']}; font-weight: bold;")
        self.sifre_input = QLineEdit()
        self.sifre_input.setEchoMode(QLineEdit.Password)
        self.sifre_input.setPlaceholderText("12345")
        self.sifre_input.returnPressed.connect(self.giris_yap)
        layout.addWidget(sifre_label)
        layout.addWidget(self.sifre_input)

        # ─── DEMO BİLGİSİ ───
        bilgi = QLabel("Demo: kutuphane / 12345")
        bilgi.setAlignment(Qt.AlignCenter)
        bilgi.setStyleSheet(f"color: {COLORS['text_sec']}; font-size: 11px; margin-top: 15px;")
        layout.addWidget(bilgi)

        layout.addSpacing(20)

        # ─── BUTONLAR ───
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)

        giris_btn = QPushButton("🔓 Giriş Yap")
        giris_btn.setMinimumHeight(45)
        giris_btn.clicked.connect(self.giris_yap)

        iptal_btn = QPushButton("❌ Çıkış")
        iptal_btn.setObjectName("danger")
        iptal_btn.setMinimumHeight(45)
        iptal_btn.clicked.connect(self.reject)

        button_layout.addWidget(giris_btn)
        button_layout.addWidget(iptal_btn)

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def giris_yap(self):
        """Giriş işlemini yap"""
        kadi = self.kadi_input.text().strip()
        sifre = self.sifre_input.text().strip()

        if not kadi or not sifre:
            QMessageBox.warning(self, "Hata", "Kullanıcı adı ve şifre giriniz!")
            return

        kullanici = self.db.kullanici_kontrol(kadi, sifre)
        if kullanici:
            self.kullanici = kullanici
            self.accept()
        else:
            QMessageBox.warning(self, "Hata", "Kullanıcı adı veya şifre yanlış!")
            self.sifre_input.clear()

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 4: DIALOG'LAR
# ═════════════════════════════════════════════════════════════════════════════

class KitapEkleDialog(QDialog):
    """Kitap Ekleme Dialog'u"""

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Yeni Kitap Ekle")
        self.setGeometry(200, 200, 500, 600)
        self.setStyleSheet(LIBRARY_STYLESHEET)
        self.result = None
        self.init_ui()

    def init_ui(self):
        """UI'yi oluştur"""
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(25, 25, 25, 25)

        # ─── BAŞLIK ───
        baslik = QLabel("📖 Yeni Kitap Ekle")
        baslik.setFont(QFont("Arial", 16, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet(f"color: {COLORS['primary']}; margin-bottom: 10px;")
        layout.addWidget(baslik)

        # ─── FORM ───
        grid = QGridLayout()
        grid.setSpacing(12)

        # Kitap Adı
        grid.addWidget(QLabel("Kitap Adı:"), 0, 0)
        self.ad_input = QLineEdit()
        self.ad_input.setPlaceholderText("Kitabın adını girin")
        grid.addWidget(self.ad_input, 0, 1)

        # Yazar
        grid.addWidget(QLabel("Yazar:"), 1, 0)
        self.yazar_input = QLineEdit()
        self.yazar_input.setPlaceholderText("Yazarın adını girin")
        grid.addWidget(self.yazar_input, 1, 1)

        # Kategori
        grid.addWidget(QLabel("Kategori:"), 2, 0)
        self.kategori_combo = QComboBox()
        self.kategori_combo.addItems(self.db.kategorileri_getir())
        grid.addWidget(self.kategori_combo, 2, 1)

        # Yayın Yılı
        grid.addWidget(QLabel("Yayın Yılı:"), 3, 0)
        self.yil_input = QSpinBox()
        self.yil_input.setMinimum(1000)
        self.yil_input.setMaximum(datetime.now().year)
        self.yil_input.setValue(2000)
        grid.addWidget(self.yil_input, 3, 1)

        # Sayfa Sayısı
        grid.addWidget(QLabel("Sayfa Sayısı:"), 4, 0)
        self.sayfa_input = QSpinBox()
        self.sayfa_input.setMinimum(1)
        self.sayfa_input.setMaximum(10000)
        grid.addWidget(self.sayfa_input, 4, 1)

        # Dil
        grid.addWidget(QLabel("Dil:"), 5, 0)
        self.dil_combo = QComboBox()
        self.dil_combo.addItems(["Türkçe", "İngilizce", "Fransızca", "Almanca", "Rusça", "Diğer"])
        grid.addWidget(self.dil_combo, 5, 1)

        # Açıklama
        grid.addWidget(QLabel("Açıklama:"), 6, 0)
        self.aciklama_input = QTextEdit()
        self.aciklama_input.setMaximumHeight(80)
        grid.addWidget(self.aciklama_input, 6, 1)

        layout.addLayout(grid)

        # ─── BUTONLAR ───
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)

        ekle_btn = QPushButton("📚 Kitap Ekle")
        ekle_btn.clicked.connect(self.ekle)

        iptal_btn = QPushButton("İptal")
        iptal_btn.setObjectName("danger")
        iptal_btn.clicked.connect(self.reject)

        button_layout.addWidget(ekle_btn)
        button_layout.addWidget(iptal_btn)
        button_layout.addStretch()

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def ekle(self):
        """Kitap ekle"""
        if not self.ad_input.text().strip() or not self.yazar_input.text().strip():
            QMessageBox.warning(self, "Uyarı", "Kitap adı ve yazar alanları zorunludur!")
            return

        self.result = (
            self.ad_input.text().strip(),
            self.yazar_input.text().strip(),
            self.kategori_combo.currentText(),
            self.yil_input.value(),
            self.sayfa_input.value(),
            self.dil_combo.currentText(),
            self.aciklama_input.toPlainText()
        )
        self.accept()


class UyeEkleDialog(QDialog):
    """Üye Ekleme Dialog'u"""

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Yeni Üye Ekle")
        self.setGeometry(200, 200, 500, 550)
        self.setStyleSheet(LIBRARY_STYLESHEET)
        self.result = None
        self.init_ui()

    def init_ui(self):
        """UI'yi oluştur"""
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(25, 25, 25, 25)

        # ─── BAŞLIK ───
        baslik = QLabel("👤 Yeni Üye Ekle")
        baslik.setFont(QFont("Arial", 16, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet(f"color: {COLORS['primary']}; margin-bottom: 10px;")
        layout.addWidget(baslik)

        # ─── FORM ───
        grid = QGridLayout()
        grid.setSpacing(12)

        # Üye No (otomatik)
        grid.addWidget(QLabel("Üye No:"), 0, 0)
        self.uye_no_input = QLineEdit(self.db.son_uye_no())
        self.uye_no_input.setReadOnly(True)
        grid.addWidget(self.uye_no_input, 0, 1)

        # Ad
        grid.addWidget(QLabel("Ad:"), 1, 0)
        self.ad_input = QLineEdit()
        grid.addWidget(self.ad_input, 1, 1)

        # Soyad
        grid.addWidget(QLabel("Soyad:"), 2, 0)
        self.soyad_input = QLineEdit()
        grid.addWidget(self.soyad_input, 2, 1)

        # Email
        grid.addWidget(QLabel("Email:"), 3, 0)
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("ornek@email.com")
        grid.addWidget(self.email_input, 3, 1)

        # Telefon
        grid.addWidget(QLabel("Telefon:"), 4, 0)
        self.tel_input = QLineEdit()
        self.tel_input.setPlaceholderText("5XX XXX XX XX")
        grid.addWidget(self.tel_input, 4, 1)

        # Adres
        grid.addWidget(QLabel("Adres:"), 5, 0)
        self.adres_input = QTextEdit()
        self.adres_input.setMaximumHeight(80)
        grid.addWidget(self.adres_input, 5, 1)

        layout.addLayout(grid)

        # ─── BUTONLAR ───
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)

        ekle_btn = QPushButton("✅ Üye Ekle")
        ekle_btn.clicked.connect(self.ekle)

        iptal_btn = QPushButton("İptal")
        iptal_btn.setObjectName("danger")
        iptal_btn.clicked.connect(self.reject)

        button_layout.addWidget(ekle_btn)
        button_layout.addWidget(iptal_btn)
        button_layout.addStretch()

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def ekle(self):
        """Üye ekle"""
        if not (self.ad_input.text().strip() and self.soyad_input.text().strip() and
                self.email_input.text().strip() and self.tel_input.text().strip()):
            QMessageBox.warning(self, "Uyarı", "Tüm zorunlu alanları doldurun!")
            return

        # Email validasyonu
        if "@" not in self.email_input.text() or "." not in self.email_input.text():
            QMessageBox.warning(self, "Uyarı", "Geçerli bir email giriniz!")
            return

        self.result = (
            self.uye_no_input.text(),
            self.ad_input.text().strip(),
            self.soyad_input.text().strip(),
            self.email_input.text().strip(),
            self.tel_input.text().strip(),
            self.adres_input.toPlainText()
        )
        self.accept()


class OduncVerDialog(QDialog):
    """Ödünç Verme Dialog'u"""

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Kitap Ödünç Ver")
        self.setGeometry(200, 200, 500, 350)
        self.setStyleSheet(LIBRARY_STYLESHEET)
        self.result = None
        self.init_ui()

    def init_ui(self):
        """UI'yi oluştur"""
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(25, 25, 25, 25)

        # ─── BAŞLIK ───
        baslik = QLabel("📚 Kitap Ödünç Ver")
        baslik.setFont(QFont("Arial", 16, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet(f"color: {COLORS['primary']}; margin-bottom: 10px;")
        layout.addWidget(baslik)

        # ─── FORM ───
        grid = QGridLayout()
        grid.setSpacing(12)

        # Kitap seç
        grid.addWidget(QLabel("Kitap:"), 0, 0)
        self.kitap_combo = QComboBox()
        mevcut_kitaplar = [k for k in self.db.kitaplari_getir() if k['durum'] == 'Mevcut']
        for kitap in mevcut_kitaplar:
            self.kitap_combo.addItem(f"📖 {kitap['ad']} - {kitap['yazar']}", kitap['kitap_id'])
        grid.addWidget(self.kitap_combo, 0, 1)

        # Üye seç
        grid.addWidget(QLabel("Üye:"), 1, 0)
        self.uye_combo = QComboBox()
        uyeler = self.db.uyeleri_getir()
        for uye in uyeler:
            self.uye_combo.addItem(f"👤 {uye['ad']} {uye['soyad']} ({uye['uye_no']})", uye['uye_id'])
        grid.addWidget(self.uye_combo, 1, 1)

        # Ödünç süresi
        grid.addWidget(QLabel("Ödünç Süresi (gün):"), 2, 0)
        self.sure_input = QSpinBox()
        self.sure_input.setMinimum(1)
        self.sure_input.setMaximum(60)
        self.sure_input.setValue(14)
        grid.addWidget(self.sure_input, 2, 1)

        layout.addLayout(grid)

        # ─── BİLGİ ───
        bilgi = QLabel("⚠️ Gecikme durumunda günlük 5 TL ücret uygulanır.")
        bilgi.setStyleSheet(f"color: {COLORS['warning']}; font-size: 11px;")
        layout.addWidget(bilgi)

        # ─── BUTONLAR ───
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)

        ver_btn = QPushButton("🔖 Ödünç Ver")
        ver_btn.clicked.connect(self.odunc_ver)

        iptal_btn = QPushButton("İptal")
        iptal_btn.setObjectName("danger")
        iptal_btn.clicked.connect(self.reject)

        button_layout.addWidget(ver_btn)
        button_layout.addWidget(iptal_btn)
        button_layout.addStretch()

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def odunc_ver(self):
        """Ödünç ver"""
        kitap_id = self.kitap_combo.currentData()
        uye_id = self.uye_combo.currentData()
        sure = self.sure_input.value()

        if kitap_id and uye_id:
            self.result = (kitap_id, uye_id, sure)
            self.accept()


class IadeAlDialog(QDialog):
    """İade Alma Dialog'u"""

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Kitap İade Al")
        self.setGeometry(200, 200, 500, 350)
        self.setStyleSheet(LIBRARY_STYLESHEET)
        self.result = None
        self.init_ui()

    def init_ui(self):
        """UI'yi oluştur"""
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(25, 25, 25, 25)

        # ─── BAŞLIK ───
        baslik = QLabel("📚 Kitap İade Al")
        baslik.setFont(QFont("Arial", 16, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet(f"color: {COLORS['primary']}; margin-bottom: 10px;")
        layout.addWidget(baslik)

        # ─── FORM ───
        grid = QGridLayout()
        grid.setSpacing(12)

        grid.addWidget(QLabel("Ödünç Kaydı:"), 0, 0)
        self.odunc_combo = QComboBox()
        odunc_list = self.db.odunc_listele(durum="Ödünçte")
        for odunc in odunc_list:
            son_tarih = datetime.fromisoformat(odunc['son_tarih'].replace(' ', 'T'))
            son_tarih_str = son_tarih.strftime("%d.%m.%Y")
            self.odunc_combo.addItem(
                f"{odunc['kitap_adi']} - {odunc['uye_adi']} {odunc['uye_soyad']} (Son: {son_tarih_str})",
                odunc['odunc_id']
            )
        grid.addWidget(self.odunc_combo, 0, 1)

        layout.addLayout(grid)

        # ─── BUTONLAR ───
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)

        iade_btn = QPushButton("✅ İade Al")
        iade_btn.clicked.connect(self.iade_al)

        iptal_btn = QPushButton("İptal")
        iptal_btn.setObjectName("danger")
        iptal_btn.clicked.connect(self.reject)

        button_layout.addWidget(iade_btn)
        button_layout.addWidget(iptal_btn)
        button_layout.addStretch()

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def iade_al(self):
        """İade al"""
        odunc_id = self.odunc_combo.currentData()
        if odunc_id:
            self.result = odunc_id
            self.accept()

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 4B: STATS WIDGET (TIER 2 — SAF PyQt5 QPainter, MATPLOTLIB YOK)
# ═════════════════════════════════════════════════════════════════════════════

class ClockWidget(QWidget):
    """Güzel analog saat widget (QPainter ile çizilmiş)"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumHeight(160)
        self.setMinimumWidth(160)
        self.setMaximumWidth(200)
        self.setMaximumHeight(200)
        self.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        
        # Timer: 1 saniyede bir güncelle
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update)
        self.timer.start(1000)

    def paintEvent(self, e):
        from PyQt5.QtGui import QPainter, QColor, QPen, QBrush, QFont
        from datetime import datetime
        import math
        
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        W, H = self.width(), self.height()
        size = min(W, H)
        
        # Merkez
        cx, cy = W / 2, H / 2
        radius = size / 2 - 8
        
        # Arka plan: koyu panel
        p.setPen(QPen(QColor(COLORS['border']), 1))
        p.setBrush(QColor(COLORS['bg_dark']))
        p.drawRoundedRect(0, 0, W, H, 12, 12)
        
        # Saat kadranı: daire
        p.setPen(QPen(QColor(COLORS['primary']), 2))
        p.setBrush(QColor(COLORS['bg_surface']))
        p.drawEllipse(int(cx - radius), int(cy - radius), int(radius * 2), int(radius * 2))
        
        # Saat çizgileri (12, 3, 6, 9)
        p.setPen(QPen(QColor(COLORS['primary']), 2))
        p.setFont(QFont("Segoe UI", 10, QFont.Bold))
        for i, label in enumerate(['12', '3', '6', '9']):
            angle = i * 90 - 90
            rad = math.radians(angle)
            x = cx + (radius - 18) * math.cos(rad)
            y = cy + (radius - 18) * math.sin(rad)
            p.drawText(int(x - 8), int(y - 8), 16, 16, Qt.AlignCenter, label)
        
        # Saat işaretleri (küçük çizgiler)
        p.setPen(QPen(QColor(COLORS['text_sec']), 1))
        for i in range(60):
            angle = i * 6 - 90
            rad = math.radians(angle)
            if i % 5 == 0:
                continue  # Büyük işaretler sadece 12,3,6,9
            x1 = cx + radius * math.cos(rad)
            y1 = cy + radius * math.sin(rad)
            x2 = cx + (radius - 4) * math.cos(rad)
            y2 = cy + (radius - 4) * math.sin(rad)
            p.drawLine(int(x1), int(y1), int(x2), int(y2))
        
        # Zamanı al
        now = datetime.now()
        h = now.hour % 12
        m = now.minute
        s = now.second
        
        # Saat iğnesi (kısa, kalın)
        h_angle = (h * 30 + m * 0.5) - 90
        h_rad = math.radians(h_angle)
        h_len = radius * 0.5
        p.setPen(QPen(QColor(COLORS['primary']), 4))
        p.drawLine(int(cx), int(cy),
                   int(cx + h_len * math.cos(h_rad)), int(cy + h_len * math.sin(h_rad)))
        
        # Dakika iğnesi (uzun, orta)
        m_angle = (m * 6 + s * 0.1) - 90
        m_rad = math.radians(m_angle)
        m_len = radius * 0.7
        p.setPen(QPen(QColor(COLORS['secondary']), 3))
        p.drawLine(int(cx), int(cy),
                   int(cx + m_len * math.cos(m_rad)), int(cy + m_len * math.sin(m_rad)))
        
        # Saniye iğnesi (uzun, ince, kırmızı)
        s_angle = (s * 6) - 90
        s_rad = math.radians(s_angle)
        s_len = radius * 0.75
        p.setPen(QPen(QColor(COLORS['danger']), 1))
        p.drawLine(int(cx), int(cy),
                   int(cx + s_len * math.cos(s_rad)), int(cy + s_len * math.sin(s_rad)))
        
        # Merkez noktası
        p.setPen(QPen(QColor(COLORS['border']), 1))
        p.setBrush(QColor(COLORS['primary']))
        p.drawEllipse(int(cx - 5), int(cy - 5), 10, 10)
        
        p.end()

    def closeEvent(self, e):
        self.timer.stop()
        super().closeEvent(e)


class _ChartBase(QWidget):
    """Ortak çizim altyapısı: kart arka planı, başlık ve boş-veri durumu."""

    def __init__(self, title="", parent=None):
        super().__init__(parent)
        self.title = title
        self.data = []
        self.colors = [COLORS['primary']]
        self.setMinimumHeight(230)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    def set_data(self, data, colors=None):
        self.data = data or []
        if colors is not None:
            self.colors = colors
        self.update()

    def _panel(self, p, W, H):
        from PyQt5.QtGui import QColor, QPen, QFont
        p.setPen(QPen(QColor(COLORS['border']), 1))
        p.setBrush(QColor(COLORS['bg_dark']))
        p.drawRoundedRect(1, 1, W - 2, H - 2, 14, 14)
        p.setPen(QColor(COLORS['text_main']))
        p.setFont(QFont("Segoe UI", 11, QFont.Bold))
        p.drawText(0, 12, W, 26, Qt.AlignHCenter | Qt.AlignVCenter, self.title)

    def _empty(self, p, W, H):
        from PyQt5.QtGui import QColor, QFont
        p.setPen(QColor(COLORS['text_sec']))
        p.setFont(QFont("Segoe UI", 10))
        p.drawText(0, 0, W, H, Qt.AlignCenter, "Veri Yok")


class QtPieChart(_ChartBase):
    """Pasta grafik (kitap durumu vb.)."""

    def __init__(self, title="", parent=None):
        super().__init__(title, parent)
        self.colors = [COLORS['success'], COLORS['danger'],
                       COLORS['primary'], COLORS['warning']]

    def paintEvent(self, e):
        from PyQt5.QtGui import QPainter, QColor, QPen, QBrush, QFont
        import math
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        W, H = self.width(), self.height()
        self._panel(p, W, H)
        toplam = sum(v for _, v in self.data) if self.data else 0
        if toplam <= 0:
            self._empty(p, W, H); p.end(); return
        cx, cy = W // 2, H // 2 + 14
        r = min(W, H - 70) // 2 - 16
        angle = -90.0
        for i, (lbl, val) in enumerate(self.data):
            span = 360.0 * val / toplam
            clr = self.colors[i % len(self.colors)]
            p.setBrush(QBrush(QColor(clr)))
            p.setPen(QPen(QColor(COLORS['bg_dark']), 2))
            p.drawPie(cx - r, cy - r, 2 * r, 2 * r, int(angle * 16), int(span * 16))
            if span > 12:
                mid = math.radians(angle + span / 2)
                tx = cx + int(r * 0.6 * math.cos(mid))
                ty = cy + int(r * 0.6 * math.sin(mid))
                p.setPen(QColor("white"))
                p.setFont(QFont("Segoe UI", 9, QFont.Bold))
                p.drawText(tx - 22, ty - 9, 44, 18, Qt.AlignCenter, f"{val / toplam * 100:.0f}%")
            angle += span
        # Legend
        p.setFont(QFont("Segoe UI", 9))
        ly = H - 16 - len(self.data) * 18
        for i, (lbl, val) in enumerate(self.data):
            clr = self.colors[i % len(self.colors)]
            p.setBrush(QColor(clr)); p.setPen(Qt.NoPen)
            p.drawRoundedRect(16, ly + i * 18, 11, 11, 3, 3)
            p.setPen(QColor(COLORS['text_sec']))
            p.drawText(33, ly + i * 18 - 2, 180, 16,
                       Qt.AlignVCenter | Qt.AlignLeft, f"{lbl}: {val}")
        p.end()


class QtBarChart(_ChartBase):
    """Dikey/yatay bar grafik."""

    def __init__(self, title="", horizontal=False, value_suffix="", parent=None):
        super().__init__(title, parent)
        self.horizontal = horizontal
        self.value_suffix = value_suffix
        self.colors = [COLORS['primary'], COLORS['secondary'], COLORS['success'],
                       COLORS['warning'], COLORS['gold'], COLORS['danger'],
                       COLORS['primary_h'], COLORS['success_h']]

    def paintEvent(self, e):
        from PyQt5.QtGui import QPainter, QColor, QPen, QBrush, QFont
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        W, H = self.width(), self.height()
        self._panel(p, W, H)
        if not self.data:
            self._empty(p, W, H); p.end(); return
        max_val = max((v for _, v in self.data), default=1) or 1
        n = len(self.data)
        if self.horizontal:
            pad_l, pad_r, pad_t, pad_b = 120, 56, 48, 16
            cw, ch = W - pad_l - pad_r, H - pad_t - pad_b
            bar_h = max(10, int(ch / (n * 1.6)))
            gap = (ch - n * bar_h) / max(n + 1, 1)
            p.setFont(QFont("Segoe UI", 8))
            for i, (lbl, val) in enumerate(self.data):
                bw = int(cw * val / max_val)
                y = int(pad_t + gap + i * (bar_h + gap))
                clr = self.colors[i % len(self.colors)]
                p.setBrush(QColor(clr)); p.setPen(Qt.NoPen)
                p.drawRoundedRect(pad_l, y, max(bw, 2), bar_h, 4, 4)
                p.setPen(QColor(COLORS['text_main']))
                p.drawText(pad_l + bw + 6, y, 60, bar_h,
                           Qt.AlignVCenter | Qt.AlignLeft, f"{val}{self.value_suffix}")
                p.setPen(QColor(COLORS['text_sec']))
                name = lbl if len(lbl) <= 17 else lbl[:16] + "…"
                p.drawText(6, y, pad_l - 12, bar_h,
                           Qt.AlignVCenter | Qt.AlignRight, name)
        else:
            pad_l, pad_r, pad_t, pad_b = 46, 18, 48, 40
            cw, ch = W - pad_l - pad_r, H - pad_t - pad_b
            bar_w = max(10, int(cw / (n * 1.7)))
            gap = (cw - n * bar_w) / max(n + 1, 1)
            for i, (lbl, val) in enumerate(self.data):
                bh = int(ch * val / max_val)
                x = int(pad_l + gap + i * (bar_w + gap))
                y = pad_t + ch - bh
                clr = self.colors[i % len(self.colors)]
                p.setBrush(QColor(clr)); p.setPen(Qt.NoPen)
                p.drawRoundedRect(x, y, bar_w, max(bh, 2), 4, 4)
                p.setPen(QColor(COLORS['text_main']))
                p.setFont(QFont("Segoe UI", 8, QFont.Bold))
                p.drawText(x - 12, y - 18, bar_w + 24, 16, Qt.AlignCenter, str(val))
                p.setPen(QColor(COLORS['text_sec']))
                p.setFont(QFont("Segoe UI", 7))
                name = lbl if len(lbl) <= 9 else lbl[:8] + "…"
                p.drawText(x - 16, pad_t + ch + 4, bar_w + 32, 18, Qt.AlignCenter, name)
            p.setPen(QPen(QColor(COLORS['border']), 1))
            p.drawLine(pad_l, pad_t + ch, pad_l + cw, pad_t + ch)
        p.end()


class StatsWidget(QWidget):
    """Dashboard grafikleri — tamamen saf PyQt5 (QPainter) ile çizilir."""

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(4, 4, 4, 4)
        root.setSpacing(12)
        self.pie = QtPieChart("📚 Kitap Durumu")
        self.bar_kat = QtBarChart("📂 Kategori Dağılımı")
        self.bar_pop = QtBarChart("🔥 En Çok Okunan Kitaplar", horizontal=True, value_suffix="x")
        grid = QGridLayout()
        grid.setSpacing(12)
        grid.addWidget(self.pie, 0, 0)
        grid.addWidget(self.bar_kat, 0, 1)
        grid.addWidget(self.bar_pop, 1, 0, 1, 2)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)
        grid.setRowStretch(0, 1)
        grid.setRowStretch(1, 1)
        root.addLayout(grid)

    def update_charts(self):
        # Pasta: kitap durumu
        mevcut = self.db.mevcut_kitap_sayisi()
        oduncte = self.db.oduncte_kitap_sayisi()
        self.pie.set_data([("Mevcut", mevcut), ("Ödünçte", oduncte)],
                          [COLORS['success'], COLORS['danger']])
        # Bar: kategori dağılımı
        kategoriler = self.db.kategori_dagilimi()[:8]
        self.bar_kat.set_data([(k['kategori_adi'], k['sayi']) for k in kategoriler])
        # Yatay bar: en çok okunanlar (sadece >0)
        en_cok = [b for b in self.db.en_cok_okunan_kitaplar(8)
                  if b.get('odunc_sayisi', 0) > 0]
        self.bar_pop.set_data([(b['ad'], b['odunc_sayisi']) for b in en_cok])

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 5: MAIN WINDOW
# ═════════════════════════════════════════════════════════════════════════════

class LibraryMainWindow(QMainWindow):
    """Ana Kütüphane Penceresi"""

    def __init__(self, kullanici, db):
        super().__init__()
        self.kullanici = kullanici
        self.db = db
        
        # Başlık
        self.setWindowTitle(f"📚 Dijital Kütüphane - {kullanici['ad']} {kullanici['soyad']}")
        self.setGeometry(50, 50, 1200, 800)
        self.setStyleSheet(LIBRARY_STYLESHEET)
        
        self.init_ui()
        self.yenile_veriler()

        # Otomatik yenileme (5 saniyede bir)
        self.timer = QTimer()
        self.timer.timeout.connect(self.yenile_veriler)
        self.timer.start(5000)

    def init_ui(self):
        """Pencereyi oluştur"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(15)

        # ─── BAŞLIK ───
        header = QLabel(f"📚 DİJİTAL KÜTÜPHANE — {self.kullanici['ad']} [{self.kullanici['rol']}]")
        header.setFont(QFont("Arial", 16, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet(f"color: {COLORS['primary']}; padding: 15px;")
        main_layout.addWidget(header)

        # ─── KPI KARTLARI (TIER 2) ───
        kpi_layout = QHBoxLayout()
        kpi_layout.setSpacing(15)

        self.kitap_card = self._create_kpi_card("📚 Kitaplar", "0", COLORS['primary'])
        self.uye_card = self._create_kpi_card("👥 Üyeler", "0", COLORS['success'])
        self.odunc_card = self._create_kpi_card("📖 Ödünçler", "0", COLORS['warning'])
        self.gecikme_card = self._create_kpi_card("💰 Gecikme", "0 TL", COLORS['danger'])

        kpi_layout.addWidget(self.kitap_card)
        kpi_layout.addWidget(self.uye_card)
        kpi_layout.addWidget(self.odunc_card)
        kpi_layout.addWidget(self.gecikme_card)
        main_layout.addLayout(kpi_layout)

        # ─── SEKMELER ───
        self.tabs = QTabWidget()

        self.dashboard_tab = self._create_dashboard_tab()
        self.kitap_tab = self.create_kitap_tab()
        self.uye_tab = self.create_uye_tab()
        self.odunc_tab = self.create_odunc_tab()

        self.tabs.addTab(self.dashboard_tab, "📊 Dashboard")
        self.tabs.addTab(self.kitap_tab, "📚 Kitaplar")
        self.tabs.addTab(self.uye_tab, "👥 Üyeler")
        self.tabs.addTab(self.odunc_tab, "📖 Ödünç İşlemleri")
        self.tabs.addTab(self._create_raporlar_tab(), "📄 Raporlar")
        self.tabs.addTab(self._create_anomali_tab(), "⚠️ Anomali Tespiti")
        self.tabs.addTab(self._create_subeler_tab(), "🏢 Şubeler")  # TIER 6
        self.tabs.addTab(self._create_audit_tab(), "📋 Audit Log")  # TIER 6

        main_layout.addWidget(self.tabs)

        # ─── ÇIKIŞ BUTONU ───
        footer_layout = QHBoxLayout()
        ornek_btn = QPushButton("🌱 Örnek Veri Yükle")
        ornek_btn.setObjectName("success")
        ornek_btn.setMaximumWidth(200)
        ornek_btn.clicked.connect(self.ornek_veri_yukle_ui)
        footer_layout.addWidget(ornek_btn)
        footer_layout.addStretch()
        cikis_btn = QPushButton("🚪 Çıkış")
        cikis_btn.setObjectName("danger")
        cikis_btn.setMaximumWidth(150)
        cikis_btn.clicked.connect(self.close)
        footer_layout.addWidget(cikis_btn)
        main_layout.addLayout(footer_layout)

        central_widget.setLayout(main_layout)

    def _create_kpi_card(self, title, value, color):
        """Gradient KPI kartı (kurs platformu stili)."""
        card = QFrame()
        card.setMinimumHeight(120)
        card.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 {COLORS['bg_card']}, stop:1 {COLORS['bg_surface']});
                border-radius: 16px;
                border-left: 5px solid {color};
                border-top: 1px solid rgba(255,255,255,0.05);
                border-bottom: 1px solid rgba(0,0,0,0.25);
            }}
        """)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(8)

        top = QHBoxLayout()
        title_label = QLabel(title)
        title_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        title_label.setStyleSheet(
            f"color: {COLORS['text_sec']}; background: transparent; "
            f"border: none; letter-spacing: 0.5px;"
        )
        top.addWidget(title_label)
        top.addStretch()
        layout.addLayout(top)

        value_label = QLabel(value)
        value_label.setFont(QFont("Segoe UI", 30, QFont.Bold))
        value_label.setStyleSheet(
            f"color: {color}; background: transparent; "
            f"border: none; letter-spacing: -1px;"
        )
        layout.addWidget(value_label)

        card.value_label = value_label  # güncellemeler için referans
        return card

    def _create_dashboard_tab(self):
        """Dashboard sekmesi (saf PyQt5 grafikler + scroll + clock)"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(12)

        # Başlık + Clock: yan yana
        top_layout = QHBoxLayout()
        baslik = QLabel("📊 Genel Durum")
        baslik.setFont(QFont("Segoe UI", 15, QFont.Bold))
        baslik.setStyleSheet(
            f"color: {COLORS['text_main']}; background: transparent; border: none;"
        )
        top_layout.addWidget(baslik)
        top_layout.addStretch()
        
        clock = ClockWidget(self)
        top_layout.addWidget(clock)
        layout.addLayout(top_layout)

        self.stats_widget = StatsWidget(self.db)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        scroll.setWidget(self.stats_widget)
        layout.addWidget(scroll)
        return widget

    def _create_raporlar_tab(self):
        """Raporlar sekmesi (TIER 3+5)"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)

        # ─── BUTONLAR ───
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        excel_btn = QPushButton("📊 Excel Raporu Oluştur")
        excel_btn.clicked.connect(self.excel_raporu_olustur_thread)

        csv_kitap_btn = QPushButton("📋 Kitaplar (CSV)")
        csv_kitap_btn.clicked.connect(lambda: self.csv_raporu_olustur_thread("kitaplar"))

        csv_uye_btn = QPushButton("👤 Üyeler (CSV)")
        csv_uye_btn.clicked.connect(lambda: self.csv_raporu_olustur_thread("uyeler"))

        csv_odunc_btn = QPushButton("🔖 Ödünçler (CSV)")
        csv_odunc_btn.clicked.connect(lambda: self.csv_raporu_olustur_thread("odunc"))

        button_layout.addWidget(excel_btn)
        button_layout.addWidget(csv_kitap_btn)
        button_layout.addWidget(csv_uye_btn)
        button_layout.addWidget(csv_odunc_btn)
        button_layout.addStretch()

        layout.addLayout(button_layout)

        # ─── PROGRESS BAR (TIER 5) ───
        self.rapor_progress = QProgressBar()
        self.rapor_progress.setStyleSheet(f"""
            QProgressBar {{
                border: 1px solid {COLORS['border']};
                border-radius: 5px;
                text-align: center;
                background-color: {COLORS['bg_surface']};
                color: {COLORS['text_main']};
            }}
            QProgressBar::chunk {{
                background-color: {COLORS['primary']};
                border-radius: 3px;
            }}
        """)
        self.rapor_progress.setVisible(False)
        layout.addWidget(self.rapor_progress)

        # ─── RAPORLARı GÖSTER ───
        rapor_label = QLabel("📊 Raporlar")
        rapor_label.setFont(QFont("Arial", 14, QFont.Bold))
        rapor_label.setStyleSheet(f"color: {COLORS['primary']};")
        layout.addWidget(rapor_label)

        self.rapor_text = QTextEdit()
        self.rapor_text.setReadOnly(True)
        self.rapor_text.setStyleSheet(f"""
            QTextEdit {{
                background-color: {COLORS['bg_surface']};
                color: {COLORS['text_main']};
                border: 1px solid {COLORS['border']};
                border-radius: 8px;
                padding: 10px;
                font-family: 'Courier New', monospace;
                font-size: 11px;
            }}
        """)
        layout.addWidget(self.rapor_text)

        widget.setLayout(layout)
        return widget

    def _create_anomali_tab(self):
        """Anomali Tespiti sekmesi (TIER 4)"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)

        # ─── BUTONLAR ───
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        hizli_btn = QPushButton("⚡ Hızlı İade Edebilen")
        hizli_btn.clicked.connect(self.anomali_hizli_iade)

        gecikme_btn = QPushButton("🔴 Sık Gecikme Yapanlar")
        gecikme_btn.clicked.connect(self.anomali_gecikme)

        az_btn = QPushButton("📕 Az Okunan Kitaplar")
        az_btn.clicked.connect(self.anomali_az_okunan)

        detay_btn = QPushButton("📌 Gecikme Detayları")
        detay_btn.clicked.connect(self.anomali_gecikme_detay)

        button_layout.addWidget(hizli_btn)
        button_layout.addWidget(gecikme_btn)
        button_layout.addWidget(az_btn)
        button_layout.addWidget(detay_btn)
        button_layout.addStretch()

        layout.addLayout(button_layout)

        # ─── ANOMALİ GÖRÜNÜMÜ ───
        anomali_label = QLabel("⚠️ Anomali Tespiti")
        anomali_label.setFont(QFont("Arial", 14, QFont.Bold))
        anomali_label.setStyleSheet(f"color: {COLORS['danger']};")
        layout.addWidget(anomali_label)

        self.anomali_text = QTextEdit()
        self.anomali_text.setReadOnly(True)
        self.anomali_text.setStyleSheet(f"""
            QTextEdit {{
                background-color: {COLORS['bg_surface']};
                color: {COLORS['text_main']};
                border: 2px solid {COLORS['danger']};
                border-radius: 8px;
                padding: 10px;
                font-family: 'Courier New', monospace;
                font-size: 11px;
            }}
        """)
        layout.addWidget(self.anomali_text)

        widget.setLayout(layout)
        return widget

    def _create_subeler_tab(self):
        """Şubeler yönetimi sekmesi (TIER 6)"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)

        # ─── BUTONLAR ───
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        ekle_btn = QPushButton("➕ Şube Ekle")
        ekle_btn.clicked.connect(self.sube_ekle_dialog)

        yenile_btn = QPushButton("🔄 Yenile")
        yenile_btn.clicked.connect(self.subeleri_yenile)

        button_layout.addWidget(ekle_btn)
        button_layout.addWidget(yenile_btn)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        # ─── ŞUBELER TABLOSU ───
        self.sube_table = QTableWidget()
        self.sube_table.setColumnCount(6)
        self.sube_table.setHorizontalHeaderLabels(
            ["ID", "Şube Adı", "Şehir", "Telefon", "Müdür", "Kitap Sayısı"]
        )
        self.sube_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.sube_table.setAlternatingRowColors(True)
        layout.addWidget(self.sube_table)

        widget.setLayout(layout)
        return widget

    def _create_audit_tab(self):
        """Audit log sekmesi (TIER 6)"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)

        # ─── BUTONLAR ───
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        yenile_btn = QPushButton("🔄 Yenile")
        yenile_btn.clicked.connect(self.audit_log_yenile)

        temizle_btn = QPushButton("🗑️ Eski Logları Sil")
        temizle_btn.setObjectName("danger")
        temizle_btn.clicked.connect(self.audit_log_temizle)

        button_layout.addWidget(yenile_btn)
        button_layout.addWidget(temizle_btn)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        # ─── AUDIT LOG TABLOSU ───
        self.audit_table = QTableWidget()
        self.audit_table.setColumnCount(7)
        self.audit_table.setHorizontalHeaderLabels(
            ["ID", "Kullanıcı", "İşlem", "Tablo", "Tarih", "Saat", "Detay"]
        )
        self.audit_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.audit_table.setAlternatingRowColors(True)
        layout.addWidget(self.audit_table)

        widget.setLayout(layout)
        return widget

    def ornek_veri_yukle_ui(self):
        """Footer butonu: örnek veriyi yükler (yalnızca boş veritabanına)."""
        if self.db.toplam_kitap_sayisi() > 0:
            QMessageBox.information(
                self, "Bilgi",
                "Veritabanında zaten kayıt var; örnek veri yalnızca BOŞ "
                "veritabanına yüklenir.\n\nSıfırdan yüklemek için 'kutuphaneler.db' "
                "dosyasını silip uygulamayı yeniden başlatın.")
            return
        try:
            ozet = self.db.ornek_veri_yukle()
            QMessageBox.information(
                self, "Başarılı",
                f"🌱 Örnek veri yüklendi!\n\n"
                f"📚 {ozet['kitap']} kitap\n"
                f"👥 {ozet['uye']} üye\n"
                f"🔖 {ozet['odunc']} aktif ödünç\n"
                f"💰 {ozet['gecikme']:.2f} TL gecikme")
            self.yenile_veriler()
        except Exception as e:
            QMessageBox.warning(self, "Hata", f"Örnek veri yüklenemedi:\n{e}")

    def _update_kpi_cards(self):
        """KPI kartlarını güncelle"""
        self.kitap_card.value_label.setText(str(self.db.toplam_kitap_sayisi()))
        self.uye_card.value_label.setText(str(self.db.toplam_uye_sayisi()))
        self.odunc_card.value_label.setText(str(self.db.aktif_odunc_sayisi()))
        self.gecikme_card.value_label.setText(f"{self.db.toplam_gecikme_ucreti():.2f} TL")
        self.stats_widget.update_charts()

    def create_kitap_tab(self):
        """Kitaplar Sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)

        # ─── FİLTRE ───
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Kategori:"))
        self.kategori_combo = QComboBox()
        self.kategori_combo.addItem("Tüm Kategoriler")
        self.kategori_combo.addItems(self.db.kategorileri_getir())
        self.kategori_combo.currentTextChanged.connect(self.kitaplari_yenile)
        filter_layout.addWidget(self.kategori_combo)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # ─── BUTONLAR ───
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        ekle_btn = QPushButton("➕ Kitap Ekle")
        ekle_btn.clicked.connect(self.kitap_ekle_dialog)
        
        sil_btn = QPushButton("🗑️ Sil")
        sil_btn.setObjectName("danger")
        sil_btn.clicked.connect(self.kitap_sil)
        
        yenile_btn = QPushButton("🔄 Yenile")
        yenile_btn.clicked.connect(self.kitaplari_yenile)

        button_layout.addWidget(ekle_btn)
        button_layout.addWidget(sil_btn)
        button_layout.addWidget(yenile_btn)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        # ─── TABLO ───
        self.kitap_table = QTableWidget()
        self.kitap_table.setColumnCount(7)
        self.kitap_table.setHorizontalHeaderLabels(
            ["ID", "Kitap Adı", "Yazar", "Kategori", "Yıl", "Sayfa", "Durum"]
        )
        self.kitap_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.kitap_table.setAlternatingRowColors(True)
        layout.addWidget(self.kitap_table)

        widget.setLayout(layout)
        return widget

    def create_uye_tab(self):
        """Üyeler Sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)

        # ─── BUTONLAR ───
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        ekle_btn = QPushButton("➕ Üye Ekle")
        ekle_btn.clicked.connect(self.uye_ekle_dialog)
        
        sil_btn = QPushButton("🗑️ Sil")
        sil_btn.setObjectName("danger")
        sil_btn.clicked.connect(self.uye_sil)
        
        yenile_btn = QPushButton("🔄 Yenile")
        yenile_btn.clicked.connect(self.uyeleri_yenile)

        button_layout.addWidget(ekle_btn)
        button_layout.addWidget(sil_btn)
        button_layout.addWidget(yenile_btn)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        # ─── TABLO ───
        self.uye_table = QTableWidget()
        self.uye_table.setColumnCount(7)
        self.uye_table.setHorizontalHeaderLabels(
            ["ID", "Üye No", "Ad", "Soyad", "Email", "Telefon", "Durum"]
        )
        self.uye_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.uye_table.setAlternatingRowColors(True)
        layout.addWidget(self.uye_table)

        widget.setLayout(layout)
        return widget

    def create_odunc_tab(self):
        """Ödünç İşlemleri Sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)

        # ─── DURUM FİLTRESİ ───
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Durum:"))
        self.durum_combo = QComboBox()
        self.durum_combo.addItems(["Tümü", "Ödünçte", "İade Edildi"])
        self.durum_combo.currentTextChanged.connect(self.odunc_yenile)
        filter_layout.addWidget(self.durum_combo)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # ─── BUTONLAR ───
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        ver_btn = QPushButton("🔖 Ödünç Ver")
        ver_btn.clicked.connect(self.odunc_ver_dialog)
        
        iade_btn = QPushButton("✅ İade Al")
        iade_btn.setObjectName("success")
        iade_btn.clicked.connect(self.iade_al_dialog)
        
        yenile_btn = QPushButton("🔄 Yenile")
        yenile_btn.clicked.connect(self.odunc_yenile)

        button_layout.addWidget(ver_btn)
        button_layout.addWidget(iade_btn)
        button_layout.addWidget(yenile_btn)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        # ─── TABLO ───
        self.odunc_table = QTableWidget()
        self.odunc_table.setColumnCount(9)
        self.odunc_table.setHorizontalHeaderLabels(
            ["ID", "Kitap", "Üye", "Ödünç Tarihi", "Son Tarih", "İade Tarihi", "Durum", "Gecikme", "Üye No"]
        )
        self.odunc_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.odunc_table.setAlternatingRowColors(True)
        layout.addWidget(self.odunc_table)

        widget.setLayout(layout)
        return widget

    # ─────────────────────────────────────────────────────────────────────────
    # TABLO YENİLEME METODLARI
    # ─────────────────────────────────────────────────────────────────────────

    def yenile_veriler(self):
        """Tüm verileri yenile"""
        self.kitaplari_yenile()
        self.uyeleri_yenile()
        self.odunc_yenile()
        self._update_kpi_cards()  # TIER 2

    def kitaplari_yenile(self):
        """Kitaplar tablosunu güncelle"""
        kategori = self.kategori_combo.currentText()
        filtre = None if kategori == "Tüm Kategoriler" else kategori

        kitaplar = self.db.kitaplari_getir(filtre)
        self.kitap_table.setRowCount(0)

        for kitap in kitaplar:
            row = self.kitap_table.rowCount()
            self.kitap_table.insertRow(row)

            self.kitap_table.setItem(row, 0, QTableWidgetItem(str(kitap['kitap_id'])))
            self.kitap_table.setItem(row, 1, QTableWidgetItem(kitap['ad']))
            self.kitap_table.setItem(row, 2, QTableWidgetItem(kitap['yazar']))
            self.kitap_table.setItem(row, 3, QTableWidgetItem(kitap['kategori_adi']))
            self.kitap_table.setItem(row, 4, QTableWidgetItem(str(kitap['yayin_yili'] or "-")))
            self.kitap_table.setItem(row, 5, QTableWidgetItem(str(kitap['sayfa_sayisi'] or "-")))

            durum_item = QTableWidgetItem(kitap['durum'])
            if kitap['durum'] == 'Mevcut':
                durum_item.setForeground(QColor(COLORS['success']))
            else:
                durum_item.setForeground(QColor(COLORS['danger']))
            self.kitap_table.setItem(row, 6, durum_item)

    def uyeleri_yenile(self):
        """Üyeler tablosunu güncelle"""
        uyeler = self.db.uyeleri_getir()
        self.uye_table.setRowCount(0)

        for uye in uyeler:
            row = self.uye_table.rowCount()
            self.uye_table.insertRow(row)

            self.uye_table.setItem(row, 0, QTableWidgetItem(str(uye['uye_id'])))
            self.uye_table.setItem(row, 1, QTableWidgetItem(uye['uye_no']))
            self.uye_table.setItem(row, 2, QTableWidgetItem(uye['ad']))
            self.uye_table.setItem(row, 3, QTableWidgetItem(uye['soyad']))
            self.uye_table.setItem(row, 4, QTableWidgetItem(uye['email']))
            self.uye_table.setItem(row, 5, QTableWidgetItem(uye['telefon']))

            durum_item = QTableWidgetItem(uye['durum'])
            if uye['durum'] == 'Aktif':
                durum_item.setForeground(QColor(COLORS['success']))
            else:
                durum_item.setForeground(QColor(COLORS['danger']))
            self.uye_table.setItem(row, 6, durum_item)

    def odunc_yenile(self):
        """Ödünç tablosunu güncelle"""
        durum = self.durum_combo.currentText()
        if durum == "Tümü":
            odunc_list = self.db.odunc_listele()
        else:
            odunc_list = self.db.odunc_listele(durum=durum)

        self.odunc_table.setRowCount(0)

        for odunc in odunc_list:
            row = self.odunc_table.rowCount()
            self.odunc_table.insertRow(row)

            self.odunc_table.setItem(row, 0, QTableWidgetItem(str(odunc['odunc_id'])))
            self.odunc_table.setItem(row, 1, QTableWidgetItem(odunc['kitap_adi']))
            self.odunc_table.setItem(row, 2, QTableWidgetItem(f"{odunc['uye_adi']} {odunc['uye_soyad']}"))

            odunc_tarihi = datetime.fromisoformat(odunc['odunc_tarihi'].replace(' ', 'T'))
            self.odunc_table.setItem(row, 3, QTableWidgetItem(odunc_tarihi.strftime("%d.%m.%Y")))

            son_tarih = datetime.fromisoformat(odunc['son_tarih'].replace(' ', 'T'))
            son_tarih_item = QTableWidgetItem(son_tarih.strftime("%d.%m.%Y"))
            if son_tarih < datetime.now() and odunc['durum'] == 'Ödünçte':
                son_tarih_item.setForeground(QColor(COLORS['danger']))
            self.odunc_table.setItem(row, 4, son_tarih_item)

            iade_tarihi = datetime.fromisoformat(odunc['iade_tarihi'].replace(' ', 'T')) if odunc['iade_tarihi'] else None
            self.odunc_table.setItem(row, 5, QTableWidgetItem(iade_tarihi.strftime("%d.%m.%Y") if iade_tarihi else "-"))

            durum_item = QTableWidgetItem(odunc['durum'])
            if odunc['durum'] == 'Ödünçte':
                durum_item.setForeground(QColor(COLORS['warning']))
            else:
                durum_item.setForeground(QColor(COLORS['success']))
            self.odunc_table.setItem(row, 6, durum_item)

            self.odunc_table.setItem(row, 7, QTableWidgetItem(f"{odunc['gecikme_ucreti']:.2f} TL"))
            self.odunc_table.setItem(row, 8, QTableWidgetItem(odunc['uye_no']))

    # ─────────────────────────────────────────────────────────────────────────
    # İŞLEM METODLARI
    # ─────────────────────────────────────────────────────────────────────────

    def kitap_ekle_dialog(self):
        """Kitap ekleme diyaloğu aç"""
        dialog = KitapEkleDialog(self.db, self)
        if dialog.exec_() == QDialog.Accepted and dialog.result:
            try:
                self.db.kitap_ekle(*dialog.result)
                QMessageBox.information(self, "Başarılı", "📚 Kitap başarıyla eklendi!")
                self.kitaplari_yenile()
            except Exception as e:
                QMessageBox.warning(self, "Hata", f"Kitap eklenirken hata: {str(e)}")

    def kitap_sil(self):
        """Seçili kitabı sil"""
        row = self.kitap_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyarı", "Silinecek bir kitap seçin!")
            return

        kitap_id = int(self.kitap_table.item(row, 0).text())
        kitap_adi = self.kitap_table.item(row, 1).text()

        reply = QMessageBox.question(
            self, "Onay", f"'{kitap_adi}' kitabını silmek istediğinize emin misiniz?"
        )

        if reply == QMessageBox.Yes:
            try:
                self.db.kitap_sil(kitap_id)
                QMessageBox.information(self, "Başarılı", "Kitap silindi!")
                self.kitaplari_yenile()
            except ValueError as e:
                QMessageBox.warning(self, "Hata", str(e))

    def uye_ekle_dialog(self):
        """Üye ekleme diyaloğu aç"""
        dialog = UyeEkleDialog(self.db, self)
        if dialog.exec_() == QDialog.Accepted and dialog.result:
            try:
                self.db.uye_ekle(*dialog.result)
                QMessageBox.information(
                    self, "Başarılı", 
                    f"👤 Üye başarıyla eklendi! No: {dialog.result[0]}"
                )
                self.uyeleri_yenile()
            except sqlite3.IntegrityError:
                QMessageBox.warning(
                    self, "Hata", 
                    "Bu email veya telefon zaten kayıtlı!"
                )

    def uye_sil(self):
        """Seçili üyeyi sil"""
        row = self.uye_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyarı", "Silinecek bir üye seçin!")
            return

        uye_id = int(self.uye_table.item(row, 0).text())
        uye_adi = f"{self.uye_table.item(row, 2).text()} {self.uye_table.item(row, 3).text()}"

        reply = QMessageBox.question(
            self, "Onay", f"'{uye_adi}' üyesini silmek istediğinize emin misiniz?"
        )

        if reply == QMessageBox.Yes:
            try:
                self.db.uye_sil(uye_id)
                QMessageBox.information(self, "Başarılı", "Üye silindi!")
                self.uyeleri_yenile()
            except ValueError as e:
                QMessageBox.warning(self, "Hata", str(e))

    def odunc_ver_dialog(self):
        """Ödünç verme diyaloğu aç"""
        mevcut_kitaplar = [k for k in self.db.kitaplari_getir() if k['durum'] == 'Mevcut']
        if not mevcut_kitaplar:
            QMessageBox.warning(self, "Uyarı", "Müsait kitap bulunamadı!")
            return

        uyeler = self.db.uyeleri_getir()
        if not uyeler:
            QMessageBox.warning(self, "Uyarı", "Kayıtlı üye bulunamadı!")
            return

        dialog = OduncVerDialog(self.db, self)
        if dialog.exec_() == QDialog.Accepted and dialog.result:
            kitap_id, uye_id, sure = dialog.result
            try:
                self.db.odunc_ver(kitap_id, uye_id, sure)
                QMessageBox.information(self, "Başarılı", "🔖 Kitap ödünç verildi!")
                self.yenile_veriler()
            except ValueError as e:
                QMessageBox.warning(self, "Hata", str(e))

    def iade_al_dialog(self):
        """İade alma diyaloğu aç"""
        odunc_list = self.db.odunc_listele(durum="Ödünçte")
        if not odunc_list:
            QMessageBox.warning(self, "Uyarı", "İade edilecek kitap bulunamadı!")
            return

        dialog = IadeAlDialog(self.db, self)
        if dialog.exec_() == QDialog.Accepted and dialog.result:
            odunc_id = dialog.result
            try:
                gecikme = self.db.iade_al(odunc_id)
                
                # Gamification: XP ekle
                odunc = next((o for o in odunc_list if o['odunc_id'] == odunc_id), None)
                if odunc:
                    self.db.uye_xp_ekle(odunc['uye_id'], 10)
                    if gecikme == 0:
                        self.db.uye_xp_ekle(odunc['uye_id'], 5)  # Bonus XP
                
                if gecikme > 0:
                    QMessageBox.information(
                        self, "İade Tamamlandı",
                        f"✅ Kitap iade alındı!\n💰 Gecikme ücreti: {gecikme:.2f} TL"
                    )
                else:
                    QMessageBox.information(self, "Başarılı", "✅ Kitap iade alındı!\n🎁 +15 XP kazandınız!")
                self.yenile_veriler()
            except ValueError as e:
                QMessageBox.warning(self, "Hata", str(e))

    def excel_raporu_olustur(self):
        """Excel raporu oluştur"""
        if not OPENPYXL_OK:
            QMessageBox.warning(self, "Hata", "openpyxl kurulu değil!\npip install openpyxl")
            return
        
        try:
            filename = self.db.excel_raporu_uret()
            rapor_text = f"""
✅ EXCEL RAPORU OLUŞTURULDU

📁 Dosya: {filename}

📊 Sayfalar:
  • Kitaplar: {self.db.toplam_kitap_sayisi()} kayıt
  • Üyeler: {self.db.toplam_uye_sayisi()} kayıt
  • Ödünçler: Tüm işlemler
  • İstatistikler: Özet bilgiler

⏰ Oluşturma Tarihi: {datetime.now().strftime("%d.%m.%Y %H:%M:%S")}
            """
            self.rapor_text.setText(rapor_text)
            QMessageBox.information(self, "Başarılı", f"Excel raporu oluşturuldu:\n{filename}")
        except Exception as e:
            QMessageBox.warning(self, "Hata", f"Rapor oluşturulamadı: {str(e)}")

    def csv_raporu_olustur(self, tip):
        """CSV raporu oluştur"""
        try:
            filename = self.db.csv_raporu_uret(tip)
            
            tip_adi = {
                "kitaplar": "Kitaplar",
                "uyeler": "Üyeler",
                "odunc": "Ödünçler"
            }.get(tip, "Bilinmeyen")
            
            rapor_text = f"""
✅ CSV RAPORU OLUŞTURULDU

📁 Dosya: {filename}
📋 Türü: {tip_adi}

⏰ Oluşturma Tarihi: {datetime.now().strftime("%d.%m.%Y %H:%M:%S")}

💡 Not: Dosya UTF-8 kodlanmıştır
            """
            self.rapor_text.setText(rapor_text)
            QMessageBox.information(self, "Başarılı", f"CSV raporu oluşturuldu:\n{filename}")
        except Exception as e:
            QMessageBox.warning(self, "Hata", f"Rapor oluşturulamadı: {str(e)}")

    def excel_raporu_olustur_thread(self):
        """Excel raporu multithreading ile oluştur (TIER 5)"""
        if not OPENPYXL_OK:
            QMessageBox.warning(self, "Hata", "openpyxl kurulu değil!\npip install openpyxl")
            return
        
        self.rapor_progress.setVisible(True)
        self.rapor_progress.setValue(0)
        
        self.report_thread = ReportWorkerThread(self.db, 'excel')
        self.report_thread.progress.connect(self.on_report_progress)
        self.report_thread.error.connect(self.on_report_error)
        self.report_thread.finished.connect(self.on_report_finished)
        self.report_thread.start()

    def csv_raporu_olustur_thread(self, tip):
        """CSV raporu multithreading ile oluştur (TIER 5)"""
        self.rapor_progress.setVisible(True)
        self.rapor_progress.setValue(0)
        
        report_type_map = {
            'kitaplar': 'csv_kitaplar',
            'uyeler': 'csv_uyeler',
            'odunc': 'csv_odunc'
        }
        
        self.report_thread = ReportWorkerThread(self.db, report_type_map[tip])
        self.report_thread.progress.connect(self.on_report_progress)
        self.report_thread.error.connect(self.on_report_error)
        self.report_thread.finished.connect(self.on_report_finished)
        self.report_thread.start()

    def on_report_progress(self, message):
        """Rapor progress güncelle"""
        self.rapor_text.setText(message)
        self.rapor_progress.setValue(min(self.rapor_progress.value() + 25, 90))

    def on_report_error(self, error):
        """Rapor hatası"""
        self.rapor_progress.setVisible(False)
        QMessageBox.warning(self, "Hata", f"Rapor oluşturulamadı: {error}")

    def on_report_finished(self):
        """Rapor tamamlandı"""
        self.rapor_progress.setValue(100)
        self.rapor_progress.setVisible(False)

    # ─────────────────────────────────────────────────────────────────────────
    # ŞUBELER YÖNETİMİ (TIER 6)
    # ─────────────────────────────────────────────────────────────────────────

    def subeleri_yenile(self):
        """Şubeler tablosunu güncelle"""
        subeler = self.db.subeleri_getir()
        self.sube_table.setRowCount(0)

        for sube in subeler:
            row = self.sube_table.rowCount()
            self.sube_table.insertRow(row)

            self.sube_table.setItem(row, 0, QTableWidgetItem(str(sube['sube_id'])))
            self.sube_table.setItem(row, 1, QTableWidgetItem(sube['sube_adi']))
            self.sube_table.setItem(row, 2, QTableWidgetItem(sube['sehir']))
            self.sube_table.setItem(row, 3, QTableWidgetItem(sube['telefon']))
            self.sube_table.setItem(row, 4, QTableWidgetItem(sube['mudur'] or "-"))
            
            # Şubeye ait kitap sayısı
            stats = self.db.sube_istatistikleri(sube['sube_id'])
            self.sube_table.setItem(row, 5, QTableWidgetItem(str(stats['toplam_kitap'])))

    def sube_ekle_dialog(self):
        """Şube ekleme diyaloğu"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Yeni Şube Ekle")
        dialog.setGeometry(300, 300, 450, 300)
        dialog.setStyleSheet(LIBRARY_STYLESHEET)

        layout = QVBoxLayout()
        layout.setSpacing(12)

        # Form alanları
        grid = QGridLayout()
        
        grid.addWidget(QLabel("Şube Adı:"), 0, 0)
        sube_adi_input = QLineEdit()
        grid.addWidget(sube_adi_input, 0, 1)

        grid.addWidget(QLabel("Şehir:"), 1, 0)
        sehir_input = QLineEdit()
        grid.addWidget(sehir_input, 1, 1)

        grid.addWidget(QLabel("Adres:"), 2, 0)
        adres_input = QTextEdit()
        adres_input.setMaximumHeight(60)
        grid.addWidget(adres_input, 2, 1)

        grid.addWidget(QLabel("Telefon:"), 3, 0)
        telefon_input = QLineEdit()
        grid.addWidget(telefon_input, 3, 1)

        grid.addWidget(QLabel("Müdür:"), 4, 0)
        mudur_input = QLineEdit()
        grid.addWidget(mudur_input, 4, 1)

        layout.addLayout(grid)

        # Butonlar
        button_layout = QHBoxLayout()
        
        def ekle():
            if not sube_adi_input.text().strip():
                QMessageBox.warning(dialog, "Uyarı", "Şube adı gerekli!")
                return
            try:
                self.db.sube_ekle(
                    sube_adi_input.text().strip(),
                    sehir_input.text().strip(),
                    adres_input.toPlainText().strip(),
                    telefon_input.text().strip(),
                    mudur_input.text().strip()
                )
                self.db.audit_log_kaydet(
                    self.kullanici['kullanici_id'],
                    'sube_ekle',
                    'subeler',
                    None
                )
                QMessageBox.information(dialog, "Başarılı", "Şube eklendi!")
                dialog.accept()
                self.subeleri_yenile()
            except Exception as e:
                QMessageBox.warning(dialog, "Hata", str(e))

        ekle_btn = QPushButton("🏢 Şube Ekle")
        ekle_btn.clicked.connect(ekle)
        
        iptal_btn = QPushButton("İptal")
        iptal_btn.setObjectName("danger")
        iptal_btn.clicked.connect(dialog.reject)

        button_layout.addWidget(ekle_btn)
        button_layout.addWidget(iptal_btn)
        button_layout.addStretch()

        layout.addLayout(button_layout)
        dialog.setLayout(layout)
        dialog.exec_()

    # ─────────────────────────────────────────────────────────────────────────
    # AUDIT LOG (TIER 6)
    # ─────────────────────────────────────────────────────────────────────────

    def audit_log_yenile(self):
        """Audit log tablosunu güncelle"""
        logs = self.db.audit_log_getir(limit=100)
        self.audit_table.setRowCount(0)

        for log in logs:
            row = self.audit_table.rowCount()
            self.audit_table.insertRow(row)

            self.audit_table.setItem(row, 0, QTableWidgetItem(str(log['log_id'])))
            self.audit_table.setItem(row, 1, QTableWidgetItem(f"{log['kullanici_adi']} {log['kullanici_soyad']}"))
            self.audit_table.setItem(row, 2, QTableWidgetItem(log['islem']))
            self.audit_table.setItem(row, 3, QTableWidgetItem(log['tablo'] or "-"))
            
            tarih = datetime.fromisoformat(log['tarih'].replace(' ', 'T'))
            self.audit_table.setItem(row, 4, QTableWidgetItem(tarih.strftime("%d.%m.%Y")))
            self.audit_table.setItem(row, 5, QTableWidgetItem(tarih.strftime("%H:%M:%S")))
            
            detay = f"ID: {log['kayit_id']}" if log['kayit_id'] else "-"
            self.audit_table.setItem(row, 6, QTableWidgetItem(detay))

    def audit_log_temizle(self):
        """Eski audit logları sil"""
        reply = QMessageBox.question(
            self, "Onay",
            "30 günden eski audit logları silmek istediğinize emin misiniz?"
        )
        if reply == QMessageBox.Yes:
            try:
                self.db.audit_log_sil_arsa(30)
                QMessageBox.information(self, "Başarılı", "Eski loglar silindi!")
                self.audit_log_yenile()
            except Exception as e:
                QMessageBox.warning(self, "Hata", str(e))

    def anomali_hizli_iade(self):
        """Hızlı iade eden üyeleri göster"""
        hizli_list = self.db.unormal_iade_hizi_tespit()
        
        if not hizli_list:
            self.anomali_text.setText("✅ Şüpheli hızlı iade etme şekli bulunamadı.")
            return
        
        text = """
⚠️ HIZLI İADE EDEN ÜYELER (Şüpheli Aktivite)

Bu üyeler 1-2 gün içinde 3+ kitap iade etmiştir.
Kitap kütüphaneden alındıktan sonra hızlıca iade edilmesi
hırsızlık şüphesini arttırır.

"""
        for user in hizli_list:
            text += f"""
👤 {user['ad']} {user['soyad']}
   Hızlı iade sayısı: {user['hizli_iade_sayisi']}
   ⚠️ Durum: Kontrol Gerekli
"""
        self.anomali_text.setText(text)

    def anomali_gecikme(self):
        """Sık gecikme yapanları göster"""
        gecikme_list = self.db.sik_gecikme_yapanlar()
        
        if not gecikme_list:
            self.anomali_text.setText("✅ Sık gecikme yapan üye bulunamadı.")
            return
        
        text = """
🔴 SIK GECİKME YAPAN ÜYELER

Bu üyeler 2+ kez kitap iadelerinde gecikmişlerdir.
Toplam gecikme ücretleri gösterilmektedir.

"""
        for user in gecikme_list:
            text += f"""
👤 {user['ad']} {user['soyad']} ({user['uye_no']})
   Gecikme sayısı: {user['gecikme_sayisi']}
   💰 Toplam ücret: {user['toplam_gecikme']:.2f} TL
"""
        self.anomali_text.setText(text)

    def anomali_az_okunan(self):
        """Az okunan kitapları göster"""
        az_list = self.db.az_okunan_kitaplar()
        
        if not az_list:
            self.anomali_text.setText("✅ Hiç ödünç alınmamış kitap bulunamadı.")
            return
        
        text = f"""
📕 AZ OKUNAN KİTAPLAR ({len(az_list)} kitap)

Hiçbir üye tarafından ödünç alınmamış kitaplar.
Bu kitaplar bağışlanabilir veya koleksiyondan çıkarılabilir.

"""
        for kitap in az_list[:20]:  # İlk 20
            text += f"""
📖 {kitap['ad']}
   Yazar: {kitap['yazar']}
"""
        if len(az_list) > 20:
            text += f"\n... ve {len(az_list) - 20} kitap daha"
        
        self.anomali_text.setText(text)

    def anomali_gecikme_detay(self):
        """Gecikme detaylarını göster"""
        detay_list = self.db.gecikme_analizi()
        
        if not detay_list:
            self.anomali_text.setText("✅ Gecikmiş kitap bulunamadı.")
            return
        
        text = f"""
📌 GECİKMİŞ KİTAPLAR ({len(detay_list)} adet)

Şu anda ödünçte olan ve son tarihini geçmiş kitaplar.
İletişim bilgileri ile birlikte gösterilmektedir.

"""
        for odunc in detay_list:
            text += f"""
📖 {odunc['kitap_adi']}
👤 {odunc['uye_adi']} {odunc['uye_soyad']}
📞 {odunc['telefon']}
⏰ Gecikme: {odunc['gecikme_gunu']} gün
💰 Ücret: {odunc['gecikme_gunu'] * 5:.2f} TL

"""
        self.anomali_text.setText(text)

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 8: MULTITHREADING (TIER 5)
# ═════════════════════════════════════════════════════════════════════════════

class ReportWorkerThread(QThread):
    """Ağır raporları arka planda çalıştır"""
    finished = pyqtSignal()
    error = pyqtSignal(str)
    progress = pyqtSignal(str)

    def __init__(self, db, report_type):
        super().__init__()
        self.db = db
        self.report_type = report_type

    def run(self):
        try:
            if self.report_type == 'excel':
                self.progress.emit("Excel raporu oluşturuluyor...")
                filename = self.db.excel_raporu_uret()
                self.progress.emit(f"✅ Excel raporu oluşturuldu: {filename}")
            
            elif self.report_type == 'csv_kitaplar':
                self.progress.emit("Kitaplar CSV'si oluşturuluyor...")
                filename = self.db.csv_raporu_uret('kitaplar')
                self.progress.emit(f"✅ CSV oluşturuldu: {filename}")
            
            elif self.report_type == 'csv_uyeler':
                self.progress.emit("Üyeler CSV'si oluşturuluyor...")
                filename = self.db.csv_raporu_uret('uyeler')
                self.progress.emit(f"✅ CSV oluşturuldu: {filename}")
            
            elif self.report_type == 'csv_odunc':
                self.progress.emit("Ödünçler CSV'si oluşturuluyor...")
                filename = self.db.csv_raporu_uret('odunc')
                self.progress.emit(f"✅ CSV oluşturuldu: {filename}")
            
            self.finished.emit()
        except Exception as e:
            self.error.emit(str(e))

# ═════════════════════════════════════════════════════════════════════════════
# MAIN & APPLICATION START (TIER 1-6 Final)
# ═════════════════════════════════════════════════════════════════════════════

def main():
    """Uygulamayı başlat"""
    app = QApplication(sys.argv)
    app.setStyle('Fusion')

    # Veritabanı oluştur
    db = DatabaseManager()

    # İlk çalıştırmada veritabanı boşsa örnek verilerle doldur
    if db.toplam_kitap_sayisi() == 0:
        ozet = db.ornek_veri_yukle()
        if ozet:
            print(f"✓ Örnek veri yüklendi: {ozet['kitap']} kitap, "
                  f"{ozet['uye']} üye, gecikme {ozet['gecikme']:.2f} TL.")

    # Giriş penceresini aç
    login = LoginDialog(db)
    if login.exec_() == QDialog.Accepted and login.kullanici:
        window = LibraryMainWindow(login.kullanici, db)
        window.show()
        sys.exit(app.exec_())
    else:
        sys.exit()

if __name__ == "__main__":
    main()